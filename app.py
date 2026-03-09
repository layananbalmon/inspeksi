from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, current_app, send_file, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
from flask_login import UserMixin
from datetime import datetime
import os
import json
import uuid
from functools import wraps
import pandas as pd  
from openpyxl import Workbook, load_workbook  
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side  
from io import BytesIO  
from sqlalchemy import extract
import cloudinary
import cloudinary.uploader
import cloudinary.api
from cloudinary.utils import cloudinary_url
from dotenv import load_dotenv
import urllib.parse
from sqlalchemy import create_engine
from sqlalchemy.pool import NullPool

# Load environment variables
load_dotenv()

# ============================================================================
# KONFIGURASI APLIKASI
# ============================================================================

app = Flask(__name__)
app.secret_key = 'kunci-rahasia-stasiun-monitoring-2026'

# ============================================================================
# KONFIGURASI DATABASE (SQLITE UNTUK LOKAL, POSTGRESQL UNTUK PRODUCTION)
# ============================================================================

basedir = os.path.abspath(os.path.dirname(__file__))

# Deteksi environment
is_production = os.environ.get('RENDER') or os.environ.get('DATABASE_URL') is not None

if is_production:
    # PRODUCTION: Pakai PostgreSQL dari environment variable
    database_url = os.environ.get('DATABASE_URL')
    
    # Fix untuk Supabase (handle password dengan karakter khusus)
    if database_url and database_url.startswith('postgresql://'):
        # Supabase sudah memberikan URL yang benar
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
        print("✅ Using PostgreSQL database (Supabase)")
    else:
        # Fallback ke SQLite jika DATABASE_URL tidak valid
        app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "db_opsel.db")}'
        print("⚠️ DATABASE_URL tidak valid, menggunakan SQLite")
else:
    # DEVELOPMENT: Pakai SQLite lokal
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(basedir, "db_opsel.db")}'
    print("✅ Using SQLite database (local development)")

# Konfigurasi tambahan untuk production
if is_production:
    # Pool settings untuk PostgreSQL
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_size': 5,
        'pool_recycle': 300,
        'pool_pre_ping': True,
    }

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

app.config['SESSION_COOKIE_SECURE'] = False  # Set True jika pakai HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # Session lifetime 1 jam

# Custom flash categories dengan CSS class
FLASH_CATEGORIES = {
    'success': 'alert-success',
    'error': 'alert-danger',
    'warning': 'alert-warning',
    'info': 'alert-info',
    'debug': 'alert-secondary'
}

# ============================================================================
# KONFIGURASI CLOUDINARY
# ============================================================================

# Konfigurasi Cloudinary dari environment variables
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key=os.environ.get('CLOUDINARY_API_KEY'),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET'),
    secure=True  # Gunakan HTTPS
)

print("=" * 60)
print("✅ Cloudinary configured")
print(f"   Cloud Name: {os.environ.get('CLOUDINARY_CLOUD_NAME', 'NOT SET')}")
print("=" * 60)

# ============================================================================
# HEADER CACHE CONTROL UNTUK MENCEGAH CACHE SETELAH LOGOUT
# ============================================================================

@app.after_request
def add_header(response):
    """
    Add headers to both force latest IE rendering engine or Chrome Frame,
    and also to cache the rendered page for 10 minutes.
    """
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# Daftar operator dan kota
OPERATORS = ['telkom', 'telkomsel', 'indosat', 'xl']
KOTA_LIST = [
    'Kota Samarinda', 'Kota Balikpapan', 'Kota Bontang', 'Kutai Barat',
    'Kutai Kartanegara', 'Kutai Timur', 'Berau', 'Paser', 'Penajam Paser Utara', 'Mahakam Ulu'
]
STATUS_OPTIONS = {
    'aktif': 'Aktif',
    'tidak_aktif': 'Tidak Aktif',
    'tidak_berizin': 'Tidak Berizin',
    'tidak_sesuai': 'Tidak Sesuai Parameter'
}

# ============================================================================
# JSON ENCODER
# ============================================================================
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        # Handle UploadGambar objects
        if isinstance(obj, UploadGambar):
            thumbnail_url = None
            if obj.public_id:
                thumbnail_url, _ = cloudinary_url(
                    obj.public_id, 
                    width=200, 
                    height=200, 
                    crop="fill", 
                    quality="auto"
                )
            
            return {
                'id': obj.id,
                'stasiun_id': obj.stasiun_id,
                'stasiun_lawan_id': obj.stasiun_lawan_id,
                'group_id': obj.group_id,
                'public_id': obj.public_id,
                'url': obj.cloudinary_url,
                'thumbnail_url': thumbnail_url,
                'original_filename': obj.original_filename,
                'status': obj.status,
                'uploaded_by': obj.uploaded_by,
                'uploaded_at': obj.uploaded_at.isoformat() if obj.uploaded_at else None,
                'is_checked': obj.is_checked,
                'width': obj.width,
                'height': obj.height,
                'format': obj.format,
                'bytes_size': obj.bytes_size
            }
        # Handle StasiunLawan objects
        elif isinstance(obj, StasiunLawan):
            return {
                'id': obj.id,
                'nama_stasiun_lawan': obj.nama_stasiun_lawan,
                'stasiun_id': obj.stasiun_id,
                'freq_tx': obj.freq_tx,
                'freq_rx': obj.freq_rx,
                'group_id': obj.group_id,
                'urutan': obj.urutan,
                'created_at': obj.created_at.isoformat() if obj.created_at else None
            }
        # Handle datetime
        elif isinstance(obj, datetime):
            return obj.isoformat()
        # Handle SQLAlchemy models
        elif hasattr(obj, '__table__'):
            # For SQLAlchemy models
            data = {}
            for column in obj.__table__.columns:
                value = getattr(obj, column.name)
                if isinstance(value, datetime):
                    data[column.name] = value.isoformat()
                else:
                    data[column.name] = value
            return data
        # Handle other objects with __dict__
        elif hasattr(obj, '__dict__'):
            try:
                # Remove internal attributes
                data = {}
                for key, value in obj.__dict__.items():
                    if not key.startswith('_'):
                        if isinstance(value, datetime):
                            data[key] = value.isoformat()
                        else:
                            data[key] = value
                return data
            except:
                pass
        return super().default(obj)

app.json_encoder = CustomJSONEncoder

# ============================================================================
# DATABASE MODELS
# ============================================================================

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu untuk mengakses halaman ini.'
login_manager.login_message_category = 'warning'
login_manager.refresh_view = 'login'
login_manager.needs_refresh_message = 'Sesi Anda telah berakhir, silakan login kembali.'
login_manager.needs_refresh_message_category = 'warning'

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), nullable=False)  # admin_master, admin_operator, user_operator
    operator_type = db.Column(db.String(50), nullable=True)  # telkom, telkomsel, indosat, xl, NULL untuk admin_master
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def __repr__(self):
        return f'<User {self.username} - {self.role} - {self.operator_type}>'

class GrupStasiun(db.Model):
    __tablename__ = 'grup_stasiun'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    stasiun_id = db.Column(db.Integer, db.ForeignKey('stasiun.id'), nullable=False)
    nama_grup = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relasi
    stasiun = db.relationship('Stasiun', backref='grup_list')
    lawan_list = db.relationship('StasiunLawan', backref='grup', lazy=True)

class Stasiun(db.Model):
    __tablename__ = 'stasiun'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    stasiun_name = db.Column(db.String(200), nullable=False)
    operator = db.Column(db.String(50), nullable=False)
    kota = db.Column(db.String(100), nullable=False)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    stasiun_lawan_list = db.relationship('StasiunLawan', backref='stasiun', lazy=True, cascade='all, delete-orphan')
    uploads = db.relationship('UploadGambar', backref='stasiun', lazy=True, cascade='all, delete-orphan')
    
    def __repr__(self):
        return f'<Stasiun {self.stasiun_name} - {self.operator}>'

class StasiunLawan(db.Model):
    __tablename__ = 'stasiun_lawan'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    stasiun_id = db.Column(db.Integer, db.ForeignKey('stasiun.id'), nullable=False)
    nama_stasiun_lawan = db.Column(db.String(200), nullable=False)
    freq_tx = db.Column(db.String(50), nullable=True) 
    freq_rx = db.Column(db.String(50), nullable=True) 
    group_id = db.Column(db.Integer, nullable=True)
    grup_id = db.Column(db.Integer, db.ForeignKey('grup_stasiun.id'), nullable=True)
    urutan = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    uploads = db.relationship('UploadGambar', backref='stasiun_lawan_obj', lazy=True)
    status_updates = db.relationship('StatusUpdate', backref='stasiun_lawan', lazy=True, cascade='all, delete-orphan')
    
class UploadGambar(db.Model):
    __tablename__ = 'upload_gambar'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    stasiun_id = db.Column(db.Integer, db.ForeignKey('stasiun.id'), nullable=False)
    stasiun_lawan_id = db.Column(db.Integer, db.ForeignKey('stasiun_lawan.id'), nullable=True)
    group_id = db.Column(db.Integer, nullable=True)
    
    # ===== CLOUDINARY FIELDS =====
    public_id = db.Column(db.String(300), nullable=False)  # Cloudinary public ID
    cloudinary_url = db.Column(db.String(500), nullable=False)  # URL dari Cloudinary
    original_filename = db.Column(db.String(300), nullable=True)  # Nama file asli
    
    # Metadata dari Cloudinary
    width = db.Column(db.Integer, nullable=True)
    height = db.Column(db.Integer, nullable=True)
    format = db.Column(db.String(10), nullable=True)
    bytes_size = db.Column(db.Integer, nullable=True)  # Ukuran file dalam bytes
    # =================================
    
    status = db.Column(db.String(50), nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_checked = db.Column(db.Boolean, default=False)
    
    def __repr__(self):
        return f'<UploadGambar {self.public_id}>'
    
    def to_dict(self):
        """Untuk serialisasi JSON"""
        # Dapatkan URL thumbnail dari Cloudinary
        thumbnail_url = None
        if self.public_id:
            thumbnail_url, _ = cloudinary_url(
                self.public_id, 
                width=200, 
                height=200, 
                crop="fill", 
                quality="auto"
            )
        
        return {
            'id': self.id,
            'stasiun_id': self.stasiun_id,
            'stasiun_lawan_id': self.stasiun_lawan_id,
            'group_id': self.group_id,
            'public_id': self.public_id,
            'url': self.cloudinary_url,
            'thumbnail_url': thumbnail_url,
            'original_filename': self.original_filename,
            'status': self.status,
            'uploaded_by': self.uploaded_by,
            'uploaded_at': self.uploaded_at.isoformat() if self.uploaded_at else None,
            'is_checked': self.is_checked,
            'width': self.width,
            'height': self.height,
            'format': self.format,
            'size_mb': round(self.bytes_size / (1024 * 1024), 2) if self.bytes_size else None
        }
    
class StatusUpdate(db.Model):
    __tablename__ = 'status_updates'
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    stasiun_lawan_id = db.Column(db.Integer, db.ForeignKey('stasiun_lawan.id'), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    updated_by = db.Column(db.Integer, db.ForeignKey('users.id'))
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)
    catatan = db.Column(db.Text, nullable=True)
    
    def __repr__(self):
        return f'<StatusUpdate {self.status}>'

@login_manager.user_loader
def load_user(user_id):
    # FIX: Menggunakan SQLAlchemy 2.0 style untuk menghindari warning
    try:
        return db.session.get(User, int(user_id))
    except:
        return None

# ============================================================================
# CONTEXT PROCESSORS
# ============================================================================

@app.context_processor
def inject_models():
    """Inject models ke semua template"""
    return dict(
        StatusUpdate=StatusUpdate,
        db=db,
        Stasiun=Stasiun,
        StasiunLawan=StasiunLawan,
        UploadGambar=UploadGambar,
        User=User
    )

# ============================================================================
# CUSTOM DECORATORS
# ============================================================================

def admin_master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'admin_master':
            flash('Akses ditolak! Hanya untuk Admin Master.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def admin_operator_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'admin_operator':
            flash('Akses ditolak! Hanya untuk Admin Operator.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def user_operator_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role != 'user_operator':
            flash('Akses ditolak! Hanya untuk User Operator.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def admin_master_or_admin_operator_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.role not in ['admin_master', 'admin_operator']:
            flash('Akses ditolak! Hanya untuk Admin.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def get_operator_filter():
    """Get operator filter based on user role"""
    if current_user.role == 'admin_master':
        return OPERATORS
    elif current_user.role == 'admin_operator':
        return [current_user.operator_type]
    elif current_user.role == 'user_operator':
        return [current_user.operator_type]
    return []

def get_accessible_operators():
    """Get list of operators accessible by current user"""
    if current_user.role == 'admin_master':
        return OPERATORS
    elif current_user.role == 'admin_operator':
        return [current_user.operator_type]
    elif current_user.role == 'user_operator':
        return [current_user.operator_type]
    return []

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def create_upload_folders():
    """Membuat folder upload jika belum ada"""
    upload_folder = app.config['UPLOAD_FOLDER']
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)
        for operator in OPERATORS:
            operator_folder = os.path.join(upload_folder, operator)
            if not os.path.exists(operator_folder):
                os.makedirs(operator_folder)
                # Buat subfolder untuk stasiun-stasiun
                for i in range(1, 11):  # Contoh: buat 10 folder awal
                    station_folder = os.path.join(operator_folder, str(i))
                    os.makedirs(station_folder, exist_ok=True)

def sanitize_filename(filename):
    """Bersihkan nama file dari karakter tidak valid"""
    filename = secure_filename(filename)
    filename = filename.replace('@', '_').replace(' ', '_')
    return filename

def generate_unique_filename(original_filename, station_id, opponent_id=None, group_id=None):
    """
    Generate nama file yang UNIK untuk menghindari konflik antar user
    Format: YYYYMMDD_HHMMSS_FFF_RANDOM_NAMAAWAL.EXT
    Contoh: 20260226_103045_123_a1b2c3_screenshoot1.jpg
    """
    if not original_filename or '.' not in original_filename:
        return None
    
    # Bersihkan nama file asli
    safe_name = sanitize_filename(original_filename)
    
    # Ambil ekstensi file
    name_part, ext_part = safe_name.rsplit('.', 1)
    
    # Generate timestamp dengan microsecond
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')[:-3]
    
    # Generate random string (6 karakter) untuk memastikan keunikan
    random_str = uuid.uuid4().hex[:6]
    
    # Format nama file: TIMESTAMP_RANDOM_NAMAASLI.EXT
    # Contoh: 20260226_103045_123_a1b2c3_screenshoot1.jpg
    new_filename = f"{timestamp}_{random_str}_{safe_name}"
    
    # Potong jika terlalu panjang (maks 200 karakter)
    if len(new_filename) > 200:
        # Potong nama file asli, pertahankan timestamp dan random
        max_name_length = 200 - len(timestamp) - len(random_str) - 10
        safe_name = safe_name[:max_name_length]
        new_filename = f"{timestamp}_{random_str}_{safe_name}"
    
    return new_filename

def get_stations_with_detailed_groups_paginated(operator, kota='all', search_stasiun='', page=1, per_page=20):
    """Get stations with detailed group information for user operator with pagination"""
    query = Stasiun.query.filter_by(operator=operator)
    
    # Filter kota
    if kota != 'all' and kota:
        query = query.filter_by(kota=kota)
    
    # Filter pencarian stasiun name
    if search_stasiun:
        query = query.filter(Stasiun.stasiun_name.ilike(f'%{search_stasiun}%'))
    
    # Query dengan pagination
    stations_query = query.order_by(Stasiun.stasiun_name)
    pagination = stations_query.paginate(page=page, per_page=per_page, error_out=False)
    stations = pagination.items
    
    # Ambil semua grup untuk setiap stasiun
    for station in stations:
        station.stasiun_lawan_list = StasiunLawan.query\
            .filter_by(stasiun_id=station.id)\
            .order_by(StasiunLawan.urutan)\
            .all()
        
        # Ambil semua grup untuk stasiun ini
        all_groups = GrupStasiun.query.filter_by(stasiun_id=station.id).all()
        groups_dict = {g.id: g.nama_grup for g in all_groups}
        
        # Kelompokkan lawan berdasarkan grup
        lawan_by_group = {}
        group_stats = {}
        
        for lawan in station.stasiun_lawan_list:
            group_key = lawan.group_id if lawan.group_id is not None else 'ungrouped'
            if group_key not in lawan_by_group:
                lawan_by_group[group_key] = []
                group_stats[group_key] = {
                    'total': 0,
                    'aktif': 0,
                    'tidak_aktif': 0,
                    'tidak_berizin': 0,
                    'tidak_sesuai': 0,
                    'belum_ada': 0
                }
            
            latest_status = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=lawan.id)\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
            
            status = latest_status.status if latest_status else 'belum_ada'
            status_display = STATUS_OPTIONS.get(status, 'Belum Ada') if latest_status else 'Belum Ada'
            
            lawan_data = {
                'id': lawan.id,
                'nama': lawan.nama_stasiun_lawan,
                'freq_tx': lawan.freq_tx,
                'freq_rx': lawan.freq_rx,
                'group_id': lawan.group_id,
                'status': status,
                'status_display': status_display,
                'catatan': latest_status.catatan if latest_status else ''
            }
            
            lawan_by_group[group_key].append(lawan_data)
            
            # Hitung statistik grup
            group_stats[group_key]['total'] += 1
            if status in group_stats[group_key]:
                group_stats[group_key][status] += 1
        
        # Format untuk display dengan grup
        groups_display = []
        
        # Grup tanpa grup dulu
        if 'ungrouped' in lawan_by_group:
            groups_display.append({
                'group_id': None,
                'group_name': 'Tanpa Grup',
                'lawans': lawan_by_group['ungrouped'],
                'stats': group_stats['ungrouped'],
                'is_ungrouped': True,
                'badge_color': 'warning'
            })
        
        # Grup dengan ID
        for group_key, lawans in lawan_by_group.items():
            if group_key != 'ungrouped':
                try:
                    group_id = int(group_key)
                    group_name = groups_dict.get(group_id, f'Grup {group_id}')
                    
                    groups_display.append({
                        'group_id': group_id,
                        'group_name': group_name,
                        'lawans': lawans,
                        'stats': group_stats[group_key],
                        'is_ungrouped': False,
                        'badge_color': 'success'
                    })
                except:
                    continue
        
        # Urutkan grup
        groups_display.sort(key=lambda x: (x['is_ungrouped'], x['group_id'] if x['group_id'] is not None else 0))
        
        station.groups_display = groups_display
        
        # Hitung statistik total
        station.total_opponents = len(station.stasiun_lawan_list)
        station.total_groups = len([g for g in groups_display if not g['is_ungrouped']])
        
        # Hitung jumlah upload unik - UPDATE untuk Cloudinary
        station.upload_count = db.session.query(db.func.count(db.distinct(UploadGambar.public_id)))\
            .filter_by(stasiun_id=station.id)\
            .scalar() or 0
    
    return stations, pagination

def get_grouped_opponents(station_id):
    """Get opponents grouped by group_id"""
    opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).order_by(StasiunLawan.urutan).all()
    
    # Kelompokkan berdasarkan group_id
    groups = {}
    for opponent in opponents:
        group_key = opponent.group_id if opponent.group_id is not None else 'ungrouped'
        if group_key not in groups:
            groups[group_key] = []
        groups[group_key].append(opponent)
    
    return groups

def get_grouped_opponents_serializable(station_id):
    """Get opponents grouped by group_id in serializable format"""
    opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).order_by(StasiunLawan.urutan).all()
    
    # Kelompokkan berdasarkan group_id
    groups = {}
    for opponent in opponents:
        group_key = opponent.group_id if opponent.group_id is not None else 'ungrouped'
        if group_key not in groups:
            groups[group_key] = []
        
        # Get latest status
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=opponent.id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        # Buat format serializable
        opponent_data = opponent.to_dict()
        
        if latest_status:
            opponent_data['latest_status'] = {
                'status': latest_status.status,
                'catatan': latest_status.catatan,
                'updated_at': latest_status.updated_at.isoformat() if latest_status.updated_at else None
            }
        else:
            opponent_data['latest_status'] = None
        
        groups[group_key].append(opponent_data)
    
    return groups

def get_upload_data_by_group(station_id):
    """Get upload data grouped by group_id - UPDATE untuk Cloudinary"""
    # Get all uploads for this station
    all_uploads = UploadGambar.query.filter_by(stasiun_id=station_id).all()
    
    # Get all opponents for this station
    opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).all()
    
    # Group opponents by group_id
    opponents_by_group = {}
    for opponent in opponents:
        group_key = opponent.group_id if opponent.group_id is not None else 'ungrouped'
        if group_key not in opponents_by_group:
            opponents_by_group[group_key] = []
        opponents_by_group[group_key].append(opponent)
    
    # Group uploads by group_id
    uploads_by_group = {}
    for upload in all_uploads:
        if upload.group_id is not None:
            group_key = upload.group_id
        elif upload.stasiun_lawan_id:
            # Get opponent's group_id
            opponent = StasiunLawan.query.get(upload.stasiun_lawan_id)
            if opponent:
                group_key = opponent.group_id if opponent.group_id is not None else 'ungrouped'
            else:
                group_key = 'ungrouped'
        else:
            group_key = 'ungrouped'
        
        if group_key not in uploads_by_group:
            uploads_by_group[group_key] = []
        
        # Don't add duplicate public_ids
        existing_public_ids = [u.public_id for u in uploads_by_group[group_key]]
        if upload.public_id not in existing_public_ids:
            uploads_by_group[group_key].append(upload)
    
    # Combine both
    grouped_data = {}
    all_group_keys = set(list(opponents_by_group.keys()) + list(uploads_by_group.keys()))
    
    for group_key in all_group_keys:
        grouped_data[group_key] = {
            'uploads': uploads_by_group.get(group_key, []),
            'opponents': opponents_by_group.get(group_key, [])
        }
    
    return grouped_data

def get_card_view_data(station):
    """Helper function untuk mendapatkan data card view - UPDATE untuk Cloudinary"""
    uploads = UploadGambar.query.filter_by(stasiun_id=station.id).all()
    
    # Get images with opponent info
    images = []
    for upload in uploads:
        opponent_name = None
        if upload.stasiun_lawan_id:
            lawan = StasiunLawan.query.get(upload.stasiun_lawan_id)
            opponent_name = lawan.nama_stasiun_lawan if lawan else None
        
        thumbnail_url = None
        if upload.public_id:
            thumbnail_url, _ = cloudinary_url(
                upload.public_id, 
                width=200, 
                height=200, 
                crop="fill", 
                quality="auto"
            )
        
        images.append({
            'id': upload.id,
            'url': upload.cloudinary_url,
            'thumbnail_url': thumbnail_url,
            'filename': upload.original_filename,
            'status': upload.status,
            'opponent': opponent_name,
            'date': upload.uploaded_at.strftime('%d/%m/%Y'),
            'format': upload.format,
            'width': upload.width,
            'height': upload.height,
            'size_mb': round(upload.bytes_size / (1024 * 1024), 2) if upload.bytes_size else None
        })
    
    # Get opponents list
    opponents = []
    for lawan in station.stasiun_lawan_list:
        opponents.append(lawan.nama_stasiun_lawan)
    
    return {
        'id': station.id,
        'name': station.stasiun_name,
        'kota': station.kota,
        'image_count': len(uploads),
        'opponent_count': len(station.stasiun_lawan_list),
        'opponents': opponents,
        'images': images,
        'created_at': station.created_at,
        'updated_at': station.updated_at
    }

def get_station_detail_data(station_id):
    """Fungsi untuk mendapatkan data detail stasiun - UPDATE untuk Cloudinary"""
    station = Stasiun.query.get_or_404(station_id)
    
    all_uploads_raw = UploadGambar.query.filter_by(stasiun_id=station_id).all()
    
    # Kelompokkan upload berdasarkan group_id dan public_id
    group_file_mapping = {}
    
    for upload in all_uploads_raw:
        group_key = upload.group_id if upload.group_id is not None else 'ungrouped'
        
        if group_key not in group_file_mapping:
            group_file_mapping[group_key] = []
        
        # Cek apakah public_id sudah ada dalam grup ini
        existing_public_ids = [u['public_id'] for u in group_file_mapping[group_key]]
        if upload.public_id not in existing_public_ids:
            group_file_mapping[group_key].append({
                'id': upload.id,
                'public_id': upload.public_id,
                'url': upload.cloudinary_url,
                'original_filename': upload.original_filename,
                'status': upload.status,
                'uploaded_at': upload.uploaded_at,
                'stasiun_lawan_id': upload.stasiun_lawan_id,
                'format': upload.format,
                'width': upload.width,
                'height': upload.height,
                'bytes_size': upload.bytes_size
            })
    
    # Buat struktur grouped_uploads
    grouped_uploads = {}
    
    for group_key, uploads_list in group_file_mapping.items():
        # Sort uploads terbaru dulu
        uploads_sorted = sorted(uploads_list, key=lambda x: x['uploaded_at'], reverse=True)
        
        # Dapatkan semua lawan dalam grup ini
        if group_key == 'ungrouped':
            # Untuk ungrouped, ambil lawan dari upload individual
            opponent_ids = set(u['stasiun_lawan_id'] for u in uploads_sorted if u['stasiun_lawan_id'])
            lawans_in_group = StasiunLawan.query.filter(StasiunLawan.id.in_(opponent_ids)).all() if opponent_ids else []
        else:
            # Untuk grup, ambil semua lawan dengan group_id yang sama
            lawans_in_group = StasiunLawan.query.filter_by(
                stasiun_id=station_id, 
                group_id=group_key
            ).all()
        
        # Hitung statistik status
        status_summary = {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
        opponent_list = []
        
        for lawan in lawans_in_group:
            latest_status = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=lawan.id)\
                .filter(StatusUpdate.status != 'riwayat')\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
            
            opponent_info = {
                'id': lawan.id,
                'nama': lawan.nama_stasiun_lawan,
                'freq_tx': lawan.freq_tx,
                'freq_rx': lawan.freq_rx,
                'latest_status': latest_status.status if latest_status else None,
                'catatan': latest_status.catatan if latest_status else None,
                'status_updated_at': latest_status.updated_at if latest_status else None
            }
            
            if opponent_info['latest_status'] in status_summary:
                status_summary[opponent_info['latest_status']] += 1
            
            opponent_list.append(opponent_info)
        
        # Urutkan lawan berdasarkan nama
        opponent_list.sort(key=lambda x: x['nama'])
        
        grouped_uploads[group_key] = {
            'group_id': group_key if group_key != 'ungrouped' else None,
            'uploads': uploads_sorted,
            'opponents': opponent_list,
            'opponent_count': len(opponent_list),
            'status_summary': status_summary
        }
    
    # Urutkan grup: ungrouped dulu, lalu grup numeric
    sorted_groups = {}
    
    if 'ungrouped' in grouped_uploads:
        sorted_groups['ungrouped'] = grouped_uploads.pop('ungrouped')
    
    # Urutkan grup numeric
    numeric_groups = {}
    for key, data in grouped_uploads.items():
        try:
            group_num = int(key)
            numeric_groups[group_num] = data
        except ValueError:
            continue
    
    for group_num in sorted(numeric_groups.keys()):
        sorted_groups[str(group_num)] = numeric_groups[group_num]
    
    # Hitung statistik total
    stasiun_lawan = StasiunLawan.query.filter_by(stasiun_id=station_id).all()
    status_counts = {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
    
    for lawan in stasiun_lawan:
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=lawan.id)\
            .filter(StatusUpdate.status != 'riwayat')\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        if latest_status and latest_status.status in status_counts:
            status_counts[latest_status.status] += 1
    
    return {
        'station': station,
        'all_uploads': all_uploads_raw,
        'grouped_uploads': sorted_groups,
        'status_counts': status_counts,
        'total_opponents': len(stasiun_lawan)
    }

def validate_excel_file(file_obj):
    """Validasi file Excel"""
    if not file_obj or not file_obj.filename:
        return False, "File tidak ditemukan"
    
    # Cek ekstensi
    if not file_obj.filename.lower().endswith(('.xlsx', '.xls')):
        return False, "Format file harus .xlsx atau .xls"
    
    # Cek ukuran (maks 10MB)
    file_obj.seek(0, os.SEEK_END)
    file_size = file_obj.tell()
    file_obj.seek(0)
    
    if file_size > 10 * 1024 * 1024:  # 10MB
        return False, "Ukuran file terlalu besar (maks 10MB)"
    
    return True, "File valid"

def clean_dataframe(df):
    """Bersihkan dataframe dari whitespace dan NaN"""
    # Replace NaN dengan string kosong
    df = df.fillna('')
    
    # Gunakan map() atau apply() sebagai ganti applymap()
    try:
        # Coba gunakan map jika hanya perlu membersihkan string
        df = df.map(lambda x: str(x).strip() if isinstance(x, str) else x)
    except:
        # Fallback ke apply untuk setiap kolom
        for col in df.columns:
            df[col] = df[col].apply(lambda x: str(x).strip() if isinstance(x, str) else x)
    
    return df

# ============================================================================
# FUNGSI BARU: UPLOAD KE CLOUDINARY
# ============================================================================

def upload_to_cloudinary(file, station_id, opponent_id=None, group_id=None):
    """
    Upload file ke Cloudinary dengan organisasi folder yang rapi
    Format folder: stasiun/[operator]/[tahun]/[bulan]/[station_id]/
    """
    if not file or not file.filename:
        return None, "File tidak ditemukan"
    
    if not allowed_file(file.filename):
        return None, "Format file tidak didukung"
    
    try:
        # Dapatkan objek stasiun untuk mendapatkan operator
        station = Stasiun.query.get(station_id)
        if not station:
            return None, "Stasiun tidak ditemukan"
        
        # Bersihkan nama file
        safe_filename = sanitize_filename(file.filename)
        name_without_ext, ext = os.path.splitext(safe_filename)
        
        # Generate timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Buat public_id dengan struktur folder yang rapi
        # Format: stasiun/[operator]/[tahun]/[bulan]/[station_id]/[timestamp]_[nama_file]
        year = datetime.now().strftime('%Y')
        month = datetime.now().strftime('%m')
        
        folder_path = f"stasiun/{station.operator}/{year}/{month}/{station_id}"
        public_id = f"{folder_path}/{timestamp}_{name_without_ext}"
        
        print(f"📤 Uploading to Cloudinary: {public_id}")
        
        # Upload ke Cloudinary dengan optimasi otomatis
        upload_result = cloudinary.uploader.upload(
            file,
            public_id=public_id,
            resource_type="image",
            
            # Optimasi otomatis
            eager=[
                {"width": 800, "height": 600, "crop": "limit", "quality": "auto:good"},
                {"width": 400, "height": 300, "crop": "fill", "quality": "auto:eco"}
            ],
            eager_async=True,
            
            # Tags untuk memudahkan pencarian
            tags=[f"station_{station_id}", f"operator_{station.operator}"],
            
            # Context menyimpan metadata
            context={
                "station_id": str(station_id),
                "station_name": station.stasiun_name,
                "operator": station.operator,
                "opponent_id": str(opponent_id) if opponent_id else "",
                "group_id": str(group_id) if group_id else "",
                "original_filename": safe_filename
            },
            
            # Response termasuk metadata
            return_response=True
        )
        
        print(f"✅ Upload successful!")
        print(f"   Public ID: {upload_result['public_id']}")
        print(f"   URL: {upload_result['secure_url']}")
        print(f"   Format: {upload_result['format']}")
        print(f"   Size: {upload_result['bytes']} bytes")
        
        return {
            'public_id': upload_result['public_id'],
            'url': upload_result['secure_url'],
            'format': upload_result['format'],
            'width': upload_result.get('width'),
            'height': upload_result.get('height'),
            'bytes': upload_result.get('bytes'),
            'original_filename': safe_filename
        }, None
        
    except Exception as e:
        print(f"❌ Error uploading to Cloudinary: {e}")
        import traceback
        traceback.print_exc()
        return None, str(e)


# ============================================================================
# FUNGSI BARU: HAPUS DARI CLOUDINARY
# ============================================================================

def delete_from_cloudinary(public_id):
    """
    Hapus file dari Cloudinary
    """
    try:
        result = cloudinary.uploader.destroy(public_id)
        
        if result.get('result') == 'ok':
            print(f"✅ Deleted from Cloudinary: {public_id}")
            return True, None
        else:
            print(f"⚠️ Failed to delete: {result}")
            return False, result.get('result')
            
    except Exception as e:
        print(f"❌ Error deleting from Cloudinary: {e}")
        return False, str(e)


# ============================================================================
# FUNGSI BARU: GET OPTIMIZED URL
# ============================================================================

def get_optimized_image_url(public_id, width=None, height=None, crop="fill", quality="auto"):
    """
    Dapatkan URL gambar yang sudah dioptimasi dari Cloudinary
    """
    options = {
        "quality": quality,
        "fetch_format": "auto",
        "secure": True
    }
    
    if width and height:
        options["width"] = width
        options["height"] = height
        options["crop"] = crop
    elif width:
        options["width"] = width
        options["crop"] = "scale"
    
    url, _ = cloudinary_url(public_id, **options)
    return url


# ============================================================================
# FUNGSI INI TIDAK DIGUNAKAN LAGI (COMMENT)
# ============================================================================

# def save_uploaded_file(file, station_id, opponent_id=None, group_id=None):
#     """TIDAK DIGUNAKAN - DIGANTI DENGAN upload_to_cloudinary"""
#     pass

# def delete_physical_file(filepath):
#     """TIDAK DIGUNAKAN - DIGANTI DENGAN delete_from_cloudinary"""
#     pass

# ============================================================================
# GENERATE TEMPLATE EXCEL UNTUK ADMIN MASTER
# ============================================================================

def generate_admin_master_template():
    """Generate template Excel untuk admin master dengan 6 kolom (FREQ TX dan FREQ RX opsional)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Hapus sheet default yang tidak digunakan
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)
    
    # Header tabel - 6 KOLOM (tanpa tanda kurung di header)
    headers = ['OPERATOR', 'STASIUN NAME', 'STASIUN LAWAN', 'FREQ TX', 'FREQ RX', 'KOTA/KAB']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    # =================== DATA CONTOH DENGAN 6 KOLOM ===================
    contoh_data = [
        # OPERATOR, STASIUN NAME, STASIUN LAWAN, FREQ TX, FREQ RX, KOTA/KAB
        ['TELKOM', 'TELKOM_STN_01', 'TELKOM_LAWAN_01', '1800 MHz', '1800 MHz', 'Kota Samarinda'],
        ['', '', 'TELKOM_LAWAN_02', '', '', ''],
        ['', '', 'TELKOM_LAWAN_03', '900 MHz', '900 MHz', ''],
        
        ['TELKOMSEL', 'TELKOMSEL_STN_01', 'TELKOMSEL_LAWAN_01', '', '', 'Kota Balikpapan'],
        ['', '', 'TELKOMSEL_LAWAN_02', '2100 MHz', '2100 MHz', ''],
        
        ['INDOSAT', 'INDOSAT_STN_01', 'INDOSAT_LAWAN_01', '', '', 'Kota Bontang'],
        ['', '', 'INDOSAT_LAWAN_02', '', '', ''],
        
        ['XL', 'XL_STN_01', 'XL_LAWAN_01', '2100 MHz', '2100 MHz', 'Kutai Kartanegara'],
        ['', '', 'XL_LAWAN_02', '', '', ''],
        ['', '', 'XL_LAWAN_03', '900 MHz', '900 MHz', ''],
    ]
    
    # Tambahkan data
    for row in contoh_data:
        ws.append(row)
    
    # =================== MERGE CELL ===================
    # Merge untuk TELKOM_STN_01 (baris 2-4, 3 lawan)
    if ws.max_row >= 4:
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)  # OPERATOR
        ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)  # STASIUN NAME
        ws.merge_cells(start_row=2, start_column=6, end_row=4, end_column=6)  # KOTA/KAB
    
    # Merge untuk TELKOMSEL_STN_01 (baris 5-6, 2 lawan)
    if ws.max_row >= 6:
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)  # OPERATOR
        ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)  # STASIUN NAME
        ws.merge_cells(start_row=5, start_column=6, end_row=6, end_column=6)  # KOTA/KAB
    
    # Merge untuk INDOSAT_STN_01 (baris 7-8, 2 lawan)
    if ws.max_row >= 8:
        ws.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)  # OPERATOR
        ws.merge_cells(start_row=7, start_column=2, end_row=8, end_column=2)  # STASIUN NAME
        ws.merge_cells(start_row=7, start_column=6, end_row=8, end_column=6)  # KOTA/KAB
    
    # Merge untuk XL_STN_01 (baris 9-11, 3 lawan)
    if ws.max_row >= 11:
        ws.merge_cells(start_row=9, start_column=1, end_row=11, end_column=1)  # OPERATOR
        ws.merge_cells(start_row=9, start_column=2, end_row=11, end_column=2)  # STASIUN NAME
        ws.merge_cells(start_row=9, start_column=6, end_row=11, end_column=6)  # KOTA/KAB
    
    # Center alignment untuk merge cells
    merge_cells = ['A2', 'B2', 'F2', 'A5', 'B5', 'F5', 'A7', 'B7', 'F7', 'A9', 'B9', 'F9']
    for cell_ref in merge_cells:
        if cell_ref in ws:
            ws[cell_ref].alignment = Alignment(vertical='center', horizontal='center')
    
    # Format lebar kolom
    ws.column_dimensions['A'].width = 15  # OPERATOR
    ws.column_dimensions['B'].width = 25  # STASIUN NAME
    ws.column_dimensions['C'].width = 25  # STASIUN LAWAN
    ws.column_dimensions['D'].width = 15  # FREQ TX
    ws.column_dimensions['E'].width = 15  # FREQ RX
    ws.column_dimensions['F'].width = 20  # KOTA/KAB
    
    # =================== SHEET INSTRUKSI ===================
    ws_inst = wb.create_sheet(title="Instruksi")
    
    instruksi = [
        ["INSTRUKSI PENGISIAN - ADMIN MASTER"],
        ["------------------------------------"],
        ["PERHATIAN: File ini untuk Admin Master (SEMUA OPERATOR)"],
        [""],
        ["FORMAT KOLOM (6 KOLOM):"],
        ["1. OPERATOR      : Nama operator (telkom/telkomsel/indosat/xl) - WAJIB"],
        ["2. STASIUN NAME  : Nama stasiun utama - WAJIB"],
        ["3. STASIUN LAWAN : Nama stasiun lawan - WAJIB"],
        ["4. FREQ TX       : Frekuensi Tx - OPSIONAL (boleh dikosongkan)"],
        ["5. FREQ RX       : Frekuensi Rx - OPSIONAL (boleh dikosongkan)"],
        ["6. KOTA/KAB      : Nama kota/kabupaten - WAJIB"],
        [""],
        ["CATATAN PENTING:"],
        ["- Kolom FREQ TX dan FREQ RX bersifat OPSIONAL - boleh dikosongkan"],
        ["- Jika tidak diisi, akan dibiarkan kosong di database"],
        ["- Bisa diisi nanti melalui halaman edit"],
        [""],
        ["CARA PENGISIAN:"],
        ["1. OPERATOR:"],
        ["   - Tulis sekali untuk setiap stasiun dengan banyak lawan"],
        ["   - Gunakan merge cell (contoh: A2:A4 untuk stasiun dengan 3 lawan)"],
        ["   - Pilihan: telkom, telkomsel, indosat, xl"],
        [""],
        ["2. STASIUN NAME:"],
        ["   - Tulis sekali untuk setiap stasiun dengan banyak lawan"],
        ["   - Gunakan merge cell seperti OPERATOR"],
        [""],
        ["3. STASIUN LAWAN:"],
        ["   - Satu baris untuk satu lawan"],
        ["   - Tidak perlu merge cell"],
        [""],
        ["4. FREQ TX  - OPSIONAL:"],
        ["   - Bisa dikosongkan jika tidak ada data"],
        ["   - Contoh format: 1800 MHz, 2100 MHz, 900 MHz, 850 MHz"],
        [""],
        ["5. FREQ RX - OPSIONAL:"],
        ["   - Bisa dikosongkan jika tidak ada data"],
        ["   - Contoh format: 1800 MHz, 2100 MHz, 900 MHz, 850 MHz"],
        [""],
        ["6. KOTA/KAB:"],
        ["   - Tulis sekali untuk setiap stasiun"],
        ["   - Gunakan merge cell seperti OPERATOR dan STASIUN NAME"],
    ]
    
    for row in instruksi:
        ws_inst.append(row)
    
    # Format untuk sheet instruksi
    ws_inst.column_dimensions['A'].width = 80
    ws_inst['A1'].font = Font(bold=True, size=14)
    ws_inst['A2'].font = Font(bold=True)
    
    # =================== SHEET DAFTAR OPERATOR ===================
    ws_op = wb.create_sheet(title="Daftar Operator")
    
    ws_op.append(["DAFTAR OPERATOR YANG TERSEDIA"])
    ws_op.append(["------------------------------"])
    ws_op.append(["Gunakan nama persis seperti di bawah:"])
    ws_op.append([""])
    
    # Daftar operator
    operator_list = ["telkom", "telkomsel", "indosat", "xl"]
    for op in operator_list:
        ws_op.append([op])
    
    ws_op.column_dimensions['A'].width = 20
    ws_op['A1'].font = Font(bold=True, size=12)
    
    # =================== SHEET DAFTAR KOTA ===================
    ws_kota = wb.create_sheet(title="Daftar Kota")
    
    ws_kota.append(["DAFTAR KOTA/KAB YANG TERSEDIA"])
    ws_kota.append(["-------------------------------"])
    ws_kota.append(["Gunakan nama persis seperti di bawah:"])
    ws_kota.append([""])
    
    # Daftar kota
    kota_list_display = [
        "Kota Samarinda", "Kota Balikpapan", "Kota Bontang", "Kutai Kartanegara", 
        "Kutai Barat", "Kutai Timur", "Penajam Paser Utara", "Paser",
        "Berau", "Mahakam Ulu"
    ]
    
    kota_list_sorted = sorted(kota_list_display, key=lambda x: x.lower())
    for kota in kota_list_sorted:
        ws_kota.append([kota])
    
    ws_kota.column_dimensions['A'].width = 30
    ws_kota['A1'].font = Font(bold=True, size=12)
    
    # Atur sheet order
    wb.active = wb['Template']
    
    # =================== SAVE FILE ===================
    output = BytesIO()
    
    # Simpan workbook
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Template_Admin_Master.xlsx'
    )

def process_excel_upload_admin_master(file_path, user, kota_default='samarinda'):
    """Process Excel upload untuk admin master dengan kolom FREQ TX dan FREQ RX opsional"""
    try:
        # Baca Excel dengan pandas
        df = pd.read_excel(file_path, dtype=str)
        
        # Validasi kolom minimal
        required_columns = ['OPERATOR', 'STASIUN NAME', 'STASIUN LAWAN']
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"Kolom '{col}' tidak ditemukan dalam file Excel")
        
        # Deteksi kolom FREQ TX, FREQ RX, dan KOTA/KAB (opsional)
        has_freq_tx_column = False
        has_freq_rx_column = False
        has_kota_column = False
        
        # Normalisasi nama kolom
        for col in df.columns:
            col_upper = str(col).upper().replace(' ', '')
            
            # Cek FREQ TX
            if any(x in col_upper for x in ['FREQTX', 'FREQ_TX', 'FREQ-TX', 'TX']):
                has_freq_tx_column = True
                print(f"DEBUG - Kolom FREQ TX ditemukan: {col}")
            
            # Cek FREQ RX
            if any(x in col_upper for x in ['FREQRX', 'FREQ_RX', 'FREQ-RX', 'RX']):
                has_freq_rx_column = True
                print(f"DEBUG - Kolom FREQ RX ditemukan: {col}")
            
            # Cek KOTA
            if any(x in col_upper for x in ['KOTA', 'KAB']):
                has_kota_column = True
                print(f"DEBUG - Kolom KOTA ditemukan: {col}")
        
        # Forward fill untuk merge cell
        df['OPERATOR'] = df['OPERATOR'].ffill()
        df['STASIUN NAME'] = df['STASIUN NAME'].ffill()
        
        # Bersihkan data
        df = clean_dataframe(df)
        
        # Filter baris yang valid
        df = df[(df['STASIUN NAME'] != '') | (df['STASIUN LAWAN'] != '')]
        
        if df.empty:
            return 0, 0, ["File Excel tidak mengandung data yang valid"]
        
        success_count = 0
        error_count = 0
        messages = []
        
        # Group by operator, stasiun, dan kota
        grouped = {}
        for idx, row in df.iterrows():
            if row['STASIUN NAME'] and row['STASIUN LAWAN']:
                # Tentukan kota yang akan digunakan
                if has_kota_column and 'KOTA/KAB' in df.columns and row['KOTA/KAB']:
                    kota = str(row['KOTA/KAB']).strip().lower()
                else:
                    kota = kota_default.lower()
                
                # Ambil freq_tx dan freq_rx (opsional - boleh kosong)
                freq_tx = None
                freq_rx = None
                
                if has_freq_tx_column:
                    # Coba dapatkan dari berbagai kemungkinan nama kolom
                    for col in df.columns:
                        col_upper = str(col).upper().replace(' ', '')
                        if any(x in col_upper for x in ['FREQTX', 'FREQ_TX', 'FREQ-TX', 'TX']):
                            if pd.notna(row[col]) and str(row[col]).strip():
                                freq_tx = str(row[col]).strip()
                                break
                
                if has_freq_rx_column:
                    for col in df.columns:
                        col_upper = str(col).upper().replace(' ', '')
                        if any(x in col_upper for x in ['FREQRX', 'FREQ_RX', 'FREQ-RX', 'RX']):
                            if pd.notna(row[col]) and str(row[col]).strip():
                                freq_rx = str(row[col]).strip()
                                break
                
                key = (row['OPERATOR'].lower(), row['STASIUN NAME'], kota)
                if key not in grouped:
                    grouped[key] = []
                grouped[key].append({
                    'nama': row['STASIUN LAWAN'],
                    'freq_tx': freq_tx,
                    'freq_rx': freq_rx
                })
        
        # Process setiap stasiun
        for (operator, stasiun_name, kota), lawan_list in grouped.items():
            try:
                # Validasi operator
                if operator not in OPERATORS:
                    error_count += 1
                    messages.append(f"Operator '{operator}' tidak valid untuk stasiun '{stasiun_name}'")
                    continue
                
                # Validasi kota
                if kota.title() not in [k.title() for k in KOTA_LIST]:
                    kota = kota_default
                    messages.append(f"Kota untuk stasiun '{stasiun_name}' tidak valid, menggunakan default: {kota_default}")
                
                # Cek duplikasi
                existing = Stasiun.query.filter_by(
                    stasiun_name=stasiun_name,
                    operator=operator
                ).first()
                
                if existing:
                    # Update stasiun lawan yang belum ada
                    existing_lawan_map = {l.nama_stasiun_lawan: l for l in existing.stasiun_lawan_list}
                    new_lawans = []
                    updated_tx = []
                    updated_rx = []
                    
                    for lawan_data in lawan_list:
                        if lawan_data['nama'] not in existing_lawan_map:
                            new_lawans.append(lawan_data)
                        else:
                            # Update freq_tx dan freq_rx hanya jika ada nilai baru
                            existing_lawan = existing_lawan_map[lawan_data['nama']]
                            
                            if lawan_data['freq_tx'] and existing_lawan.freq_tx != lawan_data['freq_tx']:
                                existing_lawan.freq_tx = lawan_data['freq_tx']
                                updated_tx.append(lawan_data['nama'])
                            
                            if lawan_data['freq_rx'] and existing_lawan.freq_rx != lawan_data['freq_rx']:
                                existing_lawan.freq_rx = lawan_data['freq_rx']
                                updated_rx.append(lawan_data['nama'])
                    
                    if new_lawans:
                        urutan = len(existing.stasiun_lawan_list)
                        for lawan_data in new_lawans:
                            new_lawan = StasiunLawan(
                                stasiun_id=existing.id,
                                nama_stasiun_lawan=lawan_data['nama'],
                                freq_tx=lawan_data['freq_tx'],  # Bisa None
                                freq_rx=lawan_data['freq_rx'],  # Bisa None
                                group_id=None,
                                urutan=urutan
                            )
                            db.session.add(new_lawan)
                            urutan += 1
                        
                        messages.append(f"Stasiun '{stasiun_name}' diperbarui: ditambah {len(new_lawans)} lawan baru")
                    
                    if updated_tx:
                        messages.append(f"Stasiun '{stasiun_name}': update Freq Tx untuk {len(updated_tx)} lawan")
                    
                    if updated_rx:
                        messages.append(f"Stasiun '{stasiun_name}': update Freq Rx untuk {len(updated_rx)} lawan")
                    
                    if not new_lawans and not updated_tx and not updated_rx:
                        messages.append(f"Stasiun '{stasiun_name}' sudah ada (dilewati)")
                    continue
                
                # Buat stasiun baru
                new_station = Stasiun(
                    stasiun_name=stasiun_name,
                    operator=operator,
                    kota=kota,
                    created_by=user.id
                )
                db.session.add(new_station)
                db.session.flush()
                
                # Tambah stasiun lawan dengan freq_tx dan freq_rx (boleh kosong)
                lawan_added = 0
                for i, lawan_data in enumerate(lawan_list):
                    if lawan_data['nama']:
                        new_lawan = StasiunLawan(
                            stasiun_id=new_station.id,
                            nama_stasiun_lawan=lawan_data['nama'],
                            freq_tx=lawan_data['freq_tx'],  # Bisa None
                            freq_rx=lawan_data['freq_rx'],  # Bisa None
                            group_id=None,
                            urutan=i
                        )
                        db.session.add(new_lawan)
                        lawan_added += 1
                
                success_count += 1
                messages.append(f"Stasiun baru '{stasiun_name}' ditambahkan dengan {lawan_added} lawan")
                
            except Exception as e:
                error_count += 1
                messages.append(f"Error pada stasiun '{stasiun_name}': {str(e)}")
                continue
        
        db.session.commit()
        
        summary = f"Upload selesai! {success_count} stasiun baru ditambahkan."
        if error_count > 0:
            summary += f" {error_count} error ditemukan."
        
        messages.insert(0, summary)
        return success_count, error_count, messages
        
    except Exception as e:
        db.session.rollback()
        return 0, 0, [f"Error processing file: {str(e)}"]

def process_excel_upload_admin_operator(file_path, user, kota_default=None):
    """Process Excel upload untuk admin operator dengan FREQ TX dan FREQ RX opsional"""
    try:
        operator = user.operator_type
        
        # Baca Excel dengan pandas
        df = pd.read_excel(file_path, dtype=str)
        
        print(f"DEBUG - Jumlah kolom: {len(df.columns)}")
        print(f"DEBUG - Nama kolom: {list(df.columns)}")
        
        # Identifikasi kolom yang ada (semua opsional kecuali STASIUN NAME dan STASIUN LAWAN)
        column_mapping = {}
        
        for i, col in enumerate(df.columns):
            col_upper = str(col).upper().strip()
            
            # Identifikasi STASIUN NAME (prioritas kolom pertama atau yang mengandung NAME)
            if i == 0 or 'STASIUN NAME' in col_upper or 'NAME' in col_upper:
                column_mapping[col] = 'STASIUN NAME'
            
            # Identifikasi STASIUN LAWAN (prioritas kolom kedua atau yang mengandung LAWAN)
            elif i == 1 or 'STASIUN LAWAN' in col_upper or 'LAWAN' in col_upper:
                column_mapping[col] = 'STASIUN LAWAN'
            
            # Identifikasi FREQ TX (opsional)
            elif any(x in col_upper for x in ['FREQ TX', 'FREQ_TX', 'FREQ-TX', 'TX']):
                column_mapping[col] = 'FREQ TX'
            
            # Identifikasi FREQ RX (opsional)
            elif any(x in col_upper for x in ['FREQ RX', 'FREQ_RX', 'FREQ-RX', 'RX']):
                column_mapping[col] = 'FREQ RX'
            
            # Identifikasi KOTA/KAB (opsional)
            elif any(x in col_upper for x in ['KOTA', 'KAB']):
                column_mapping[col] = 'KOTA/KAB'
            
            else:
                column_mapping[col] = f'COL{i}'
        
        # Rename kolom sesuai mapping
        df = df.rename(columns=column_mapping)
        print(f"DEBUG - Columns after rename: {list(df.columns)}")
        
        # Validasi kolom minimal
        if 'STASIUN NAME' not in df.columns:
            return 0, 0, ["Kolom STASIUN NAME tidak ditemukan dalam file Excel"]
        
        if 'STASIUN LAWAN' not in df.columns:
            return 0, 0, ["Kolom STASIUN LAWAN tidak ditemukan dalam file Excel"]
        
        # Forward fill untuk merge cell
        df['STASIUN NAME'] = df['STASIUN NAME'].ffill()
        
        if 'KOTA/KAB' in df.columns:
            df['KOTA/KAB'] = df['KOTA/KAB'].ffill()
        
        # Bersihkan data menggunakan fungsi clean_dataframe
        df = clean_dataframe(df)
        
        # Filter baris yang valid (stasiun lawan tidak boleh kosong)
        df = df[df['STASIUN LAWAN'] != '']
        
        if df.empty:
            return 0, 0, ["File Excel tidak mengandung data stasiun lawan yang valid"]
        
        success_count = 0
        error_count = 0
        messages = []
        
        # Group by stasiun dan kota
        grouped = {}
        for idx, row in df.iterrows():
            if row['STASIUN NAME']:
                # Tentukan kota yang akan digunakan (opsional)
                kota_used = None
                
                # 1. Cek dari kolom KOTA/KAB di Excel
                if 'KOTA/KAB' in df.columns and pd.notna(row['KOTA/KAB']) and str(row['KOTA/KAB']).strip():
                    kota_used = str(row['KOTA/KAB']).strip()
                
                # 2. Jika tidak ada di Excel, cek dari form upload
                if not kota_used and kota_default:
                    kota_used = kota_default
                
                # 3. Jika masih kosong, gunakan default 'samarinda'
                if not kota_used:
                    kota_used = 'samarinda'
                    messages.append(f"ℹ️ Kota untuk stasiun '{row['STASIUN NAME']}' tidak ditemukan, menggunakan default: samarinda")
                
                # Normalisasi kota ke lowercase
                kota_used = kota_used.lower()
                
                # Ambil freq_tx dan freq_rx (opsional - boleh kosong)
                freq_tx = None
                freq_rx = None
                
                if 'FREQ TX' in df.columns and pd.notna(row['FREQ TX']) and str(row['FREQ TX']).strip():
                    freq_tx = str(row['FREQ TX']).strip()
                
                if 'FREQ RX' in df.columns and pd.notna(row['FREQ RX']) and str(row['FREQ RX']).strip():
                    freq_rx = str(row['FREQ RX']).strip()
                
                key = (row['STASIUN NAME'], kota_used)
                if key not in grouped:
                    grouped[key] = []
                
                grouped[key].append({
                    'nama': row['STASIUN LAWAN'],
                    'freq_tx': freq_tx,
                    'freq_rx': freq_rx
                })
        
        print(f"DEBUG - Total grup stasiun: {len(grouped)}")
        
        # Process setiap stasiun
        for (stasiun_name, kota), lawan_list in grouped.items():
            try:
                # Validasi kota
                kota_valid = False
                kota_normalized = None
                
                for k in KOTA_LIST:
                    if k.lower() == kota.lower():
                        kota_normalized = k.lower()
                        kota_valid = True
                        break
                
                # Coba cari yang mirip
                if not kota_valid:
                    for k in KOTA_LIST:
                        if k.lower() in kota.lower() or kota.lower() in k.lower():
                            kota_normalized = k.lower()
                            kota_valid = True
                            messages.append(f"ℹ️ Kota '{kota}' dikoreksi menjadi '{k}' untuk stasiun '{stasiun_name}'")
                            break
                
                if not kota_valid:
                    # Jika tidak ditemukan, gunakan default 'samarinda'
                    kota_normalized = 'samarinda'
                    messages.append(f"⚠️ Kota '{kota}' tidak valid untuk stasiun '{stasiun_name}', menggunakan default: samarinda")
                
                # Cek duplikasi
                existing = Stasiun.query.filter_by(
                    stasiun_name=stasiun_name,
                    operator=operator
                ).first()
                
                if existing:
                    # Update stasiun lawan yang belum ada
                    existing_lawan_map = {l.nama_stasiun_lawan: l for l in existing.stasiun_lawan_list}
                    new_lawans = []
                    updated_tx = []
                    updated_rx = []
                    
                    for lawan_data in lawan_list:
                        if lawan_data['nama'] not in existing_lawan_map:
                            new_lawans.append(lawan_data)
                        else:
                            # Update freq_tx dan freq_rx hanya jika ada nilai baru
                            existing_lawan = existing_lawan_map[lawan_data['nama']]
                            
                            if lawan_data['freq_tx'] and existing_lawan.freq_tx != lawan_data['freq_tx']:
                                existing_lawan.freq_tx = lawan_data['freq_tx']
                                updated_tx.append(lawan_data['nama'])
                            
                            if lawan_data['freq_rx'] and existing_lawan.freq_rx != lawan_data['freq_rx']:
                                existing_lawan.freq_rx = lawan_data['freq_rx']
                                updated_rx.append(lawan_data['nama'])
                    
                    if new_lawans:
                        urutan = len(existing.stasiun_lawan_list)
                        for lawan_data in new_lawans:
                            new_lawan = StasiunLawan(
                                stasiun_id=existing.id,
                                nama_stasiun_lawan=lawan_data['nama'],
                                freq_tx=lawan_data['freq_tx'],  # Bisa None
                                freq_rx=lawan_data['freq_rx'],  # Bisa None
                                group_id=None,
                                urutan=urutan
                            )
                            db.session.add(new_lawan)
                            urutan += 1
                        
                        messages.append(f"✅ Stasiun '{stasiun_name}': ditambah {len(new_lawans)} lawan baru")
                    
                    if updated_tx:
                        messages.append(f"📡 Stasiun '{stasiun_name}': update Freq Tx untuk {len(updated_tx)} lawan")
                    
                    if updated_rx:
                        messages.append(f"📡 Stasiun '{stasiun_name}': update Freq Rx untuk {len(updated_rx)} lawan")
                    
                    if not new_lawans and not updated_tx and not updated_rx:
                        messages.append(f"ℹ️ Stasiun '{stasiun_name}' sudah ada (dilewati)")
                    
                    # Commit perubahan
                    db.session.commit()
                    success_count += 1
                    continue
                
                # Buat stasiun baru
                new_station = Stasiun(
                    stasiun_name=stasiun_name,
                    operator=operator,
                    kota=kota_normalized,
                    created_by=user.id
                )
                db.session.add(new_station)
                db.session.flush()
                
                # Tambah stasiun lawan dengan freq_tx dan freq_rx (boleh kosong)
                lawan_added = 0
                for i, lawan_data in enumerate(lawan_list):
                    if lawan_data['nama']:
                        new_lawan = StasiunLawan(
                            stasiun_id=new_station.id,
                            nama_stasiun_lawan=lawan_data['nama'],
                            freq_tx=lawan_data['freq_tx'],  # Bisa None
                            freq_rx=lawan_data['freq_rx'],  # Bisa None
                            group_id=None,
                            urutan=i
                        )
                        db.session.add(new_lawan)
                        lawan_added += 1
                
                success_count += 1
                messages.append(f"✅ Stasiun baru '{stasiun_name}' ditambahkan dengan {lawan_added} lawan")
                db.session.commit()
                
            except Exception as e:
                db.session.rollback()
                error_count += 1
                messages.append(f"❌ Error pada stasiun '{stasiun_name}': {str(e)}")
                print(f"ERROR DETAIL: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        summary = f"✅ Berhasil memproses {success_count} stasiun untuk operator {operator.upper()}!"
        if error_count > 0:
            summary += f" ⚠️ {error_count} error ditemukan."
        
        messages.insert(0, summary)
        return success_count, error_count, messages
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR FATAL: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0, 0, [f"❌ Error processing file: {str(e)}"]

def generate_admin_operator_template(operator):
    """Generate template Excel untuk admin operator dengan FREQ TX dan FREQ RX opsional"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)
    
    # Tentukan operator name
    operator_name = operator.upper()
    
    # Header tabel - 5 KOLOM (tanpa tanda kurung)
    headers = ['STASIUN NAME', 'STASIUN LAWAN', 'FREQ TX', 'FREQ RX', 'KOTA/KAB']
    ws.append(headers)
    
    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    
    # =================== DATA CONTOH ===================
    contoh_data = [
        # STASIUN NAME, STASIUN LAWAN, FREQ TX, FREQ RX, KOTA/KAB
        [f'{operator_name}_STN_01', f'{operator_name}_LAWAN_01', '1800 MHz', '1800 MHz', 'Kota Samarinda'],
        ['', f'{operator_name}_LAWAN_02', '', '', ''],
        ['', f'{operator_name}_LAWAN_03', '900 MHz', '900 MHz', ''],
        
        [f'{operator_name}_STN_02', f'{operator_name}_LAWAN_04', '', '', 'Kota Balikpapan'],
        
        [f'{operator_name}_STN_03', f'{operator_name}_LAWAN_05', '2100 MHz', '2100 MHz', 'Kota Bontang'],
        ['', f'{operator_name}_LAWAN_06', '', '', ''],
    ]
    
    for row in contoh_data:
        ws.append(row)
    
    # =================== MERGE CELL ===================
    if ws.max_row >= 3:
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)  # STASIUN NAME
        ws.merge_cells(start_row=2, start_column=5, end_row=4, end_column=5)  # KOTA/KAB
    
    if ws.max_row >= 6:
        ws.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)  # STASIUN NAME
        ws.merge_cells(start_row=5, start_column=5, end_row=6, end_column=5)  # KOTA/KAB
    
    # Center alignment untuk merge cells
    merge_cells = ['A2', 'E2', 'A5', 'E5']
    for cell_ref in merge_cells:
        if cell_ref in ws:
            ws[cell_ref].alignment = Alignment(vertical='center', horizontal='center')
    
    # Format lebar kolom
    ws.column_dimensions['A'].width = 25  # STASIUN NAME
    ws.column_dimensions['B'].width = 25  # STASIUN LAWAN
    ws.column_dimensions['C'].width = 15  # FREQ TX
    ws.column_dimensions['D'].width = 15  # FREQ RX
    ws.column_dimensions['E'].width = 20  # KOTA/KAB
    
    # =================== SHEET INSTRUKSI ===================
    ws_inst = wb.create_sheet(title="Instruksi")
    
    instruksi = [
        [f"INSTRUKSI PENGISIAN - OPERATOR {operator_name}"],
        ["----------------------------------------"],
        [f"PERHATIAN: File ini khusus untuk operator {operator_name}"],
        [""],
        ["FORMAT KOLOM:"],
        ["1. STASIUN NAME  : Nama stasiun utama Anda - WAJIB"],
        ["2. STASIUN LAWAN : Nama stasiun lawan/kompetitor - WAJIB"],
        ["3. FREQ TX       : Frekuensi Tx - OPSIONAL (boleh dikosongkan)"],
        ["4. FREQ RX       : Frekuensi Rx - OPSIONAL (boleh dikosongkan)"],
        ["5. KOTA/KAB      : Nama kota/kabupaten - WAJIB"],
        [""],
        ["CATATAN PENTING:"],
        ["- Kolom FREQ TX dan FREQ RX bersifat OPSIONAL"],
        ["- Boleh dikosongkan jika belum ada data frekuensi"],
        ["- Data frekuensi bisa diisi nanti melalui halaman edit"],
        [""],
        ["CARA PENGISIAN:"],
        ["1. STASIUN NAME:"],
        ["   - Tulis sekali untuk setiap stasiun dengan banyak lawan"],
        ["   - Gunakan merge cell jika diperlukan"],
        ["   - Contoh: " + operator_name + "_STN_01, TOWER_01"],
        [""],
        ["2. STASIUN LAWAN:"],
        ["   - Satu baris untuk satu lawan"],
        ["   - Tidak perlu merge cell"],
        ["   - Contoh: " + operator_name + "_LAWAN_01, COMPETITOR_01"],
        [""],
        ["3. FREQ TX (Opsional):"],
        ["   - Bisa dikosongkan jika tidak ada data"],
        ["   - Contoh: 1800 MHz, 2100 MHz, 900 MHz"],
        [""],
        ["4. FREQ RX (Opsional):"],
        ["   - Bisa dikosongkan jika tidak ada data"],
        ["   - Contoh: 1800 MHz, 2100 MHz, 900 MHz"],
        [""],
        ["5. KOTA/KAB:"],
        ["   - Tulis sekali untuk setiap stasiun"],
        ["   - Gunakan merge cell seperti STASIUN NAME"],
    ]
    
    for row in instruksi:
        ws_inst.append(row)
    
    ws_inst.column_dimensions['A'].width = 80
    ws_inst['A1'].font = Font(bold=True, size=14)
    
    # =================== SHEET DAFTAR KOTA ===================
    ws_kota = wb.create_sheet(title="Daftar Kota")
    
    ws_kota.append(["DAFTAR KOTA/KAB UNTUK OPERATOR " + operator_name])
    ws_kota.append(["---------------------------------"])
    ws_kota.append(["Gunakan nama persis seperti di bawah:"])
    ws_kota.append([""])
    
    kota_list_display = [
        "Kota Samarinda", "Kota Balikpapan", "Kota Bontang", "Kutai Kartanegara", 
        "Kutai Barat", "Kutai Timur", "Penajam Paser Utara", "Paser",
        "Berau", "Mahakam Ulu"
    ]
    
    for kota in sorted(kota_list_display, key=lambda x: x.lower()):
        ws_kota.append([kota])
    
    ws_kota.column_dimensions['A'].width = 30
    ws_kota['A1'].font = Font(bold=True, size=12)
    
    # Atur sheet order
    wb.active = wb['Template']
    
    # =================== SAVE FILE ===================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Template_{operator_name}.xlsx'
    )

# ============================================================================
# TEMPLATE FILTERS
# ============================================================================

@app.template_filter('format_datetime')
def format_datetime_filter(value, format='%d/%m/%Y'):
    if isinstance(value, datetime):
        return value.strftime(format)
    return value

@app.template_filter('status_badge_class')
def status_badge_class_filter(status):
    """Return CSS class untuk badge status"""
    classes = {
        'aktif': 'bg-success',
        'tidak_aktif': 'bg-danger',
        'tidak_berizin': 'bg-warning',
        'tidak_sesuai': 'bg-info',
        'riwayat': 'bg-secondary'
    }
    return classes.get(status, 'bg-secondary')

@app.template_filter('status_display')
def status_display_filter(status_key):
    """Return display text untuk status"""
    display_text = STATUS_OPTIONS.get(status_key, status_key)
    if status_key == 'riwayat':
        return 'Riwayat'
    return display_text

@app.template_filter('operator_color')
def operator_color_filter(operator):
    """Return warna untuk operator"""
    colors = {
        'telkom': '#E1251B',
        'telkomsel': '#E1251B',
        'indosat': '#FF6600',
        'xl': '#7F3F98'
    }
    return colors.get(operator, '#6c757d')

@app.template_filter('flash_category_class')
def flash_category_class_filter(category):
    """Return CSS class untuk flash message category"""
    return FLASH_CATEGORIES.get(category, 'alert-info')

@app.template_filter('flash_icon')
def flash_icon_filter(category):
    """Return icon untuk flash message category"""
    icons = {
        'success': 'bi-check-circle-fill',
        'error': 'bi-exclamation-triangle-fill',
        'warning': 'bi-exclamation-triangle-fill',
        'info': 'bi-info-circle-fill',
        'debug': 'bi-bug-fill'
    }
    return icons.get(category, 'bi-info-circle-fill')

# ============================================================================
# CLOUDINARY TEMPLATE FILTERS
# ============================================================================

@app.template_filter('cloudinary_thumbnail')
def cloudinary_thumbnail_filter(public_id, width=200, height=200):
    """Filter untuk mendapatkan thumbnail dari Cloudinary"""
    if not public_id:
        return ''
    url, _ = cloudinary_url(public_id, width=width, height=height, crop="fill", quality="auto")
    return url

@app.template_filter('cloudinary_optimized')
def cloudinary_optimized_filter(public_id, width=None):
    """Filter untuk mendapatkan URL optimized dari Cloudinary"""
    if not public_id:
        return ''
    options = {"quality": "auto", "fetch_format": "auto"}
    if width:
        options["width"] = width
        options["crop"] = "scale"
    url, _ = cloudinary_url(public_id, **options)
    return url

@app.template_filter('filesizeformat')
def filesizeformat_filter(bytes_value):
    """Format ukuran file dalam bytes ke format yang mudah dibaca"""
    if not bytes_value:
        return '0 KB'
    
    for unit in ['B', 'KB', 'MB', 'GB']:
        if bytes_value < 1024.0:
            return f"{bytes_value:.1f} {unit}"
        bytes_value /= 1024.0
    return f"{bytes_value:.1f} TB"

# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================

def validate_image_file(file_obj, max_size_mb=5, log_warnings=False):
    """Validate image file format and size"""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    
    # Check if file has extension
    if '.' not in file_obj.filename:
        if log_warnings:
            flash(f'File {file_obj.filename} tidak memiliki ekstensi!', 'warning')
        return False
    
    # Extract extension
    try:
        ext = file_obj.filename.rsplit('.', 1)[1].lower()
    except IndexError:
        if log_warnings:
            flash(f'File {file_obj.filename} format tidak valid!', 'warning')
        return False
    
    # Check extension
    if ext not in ALLOWED_EXTENSIONS:
        if log_warnings:
            flash(f'File {file_obj.filename} format tidak didukung!', 'warning')
        return False
    
    # Check file size
    file_content = file_obj.read()
    max_size_bytes = max_size_mb * 1024 * 1024
    
    if len(file_content) > max_size_bytes:
        if log_warnings:
            flash(f'File {file_obj.filename} terlalu besar (maks {max_size_mb}MB)!', 'warning')
        file_obj.seek(0)  # Reset file pointer
        return False
    
    # Reset file pointer for further processing
    file_obj.seek(0)
    return True

# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin_master':
            return redirect(url_for('admin_master_dashboard'))
        elif current_user.role == 'admin_operator':
            return redirect(url_for('admin_operator_dashboard'))
        elif current_user.role == 'user_operator':
            return redirect(url_for('user_operator_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user:
            if check_password_hash(user.password, password):
                login_user(user)
                flash('Login berhasil!', 'success')
                
                # Clear session untuk mencegah caching
                session.permanent = True
                
                if user.role == 'admin_master':
                    return redirect(url_for('admin_master_dashboard'))
                elif user.role == 'admin_operator':
                    return redirect(url_for('admin_operator_dashboard'))
                elif user.role == 'user_operator':
                    return redirect(url_for('user_operator_dashboard'))
            else:
                flash('Username atau password salah!', 'error')
        else:
            flash('Username atau password salah!', 'error')
    
    response = make_response(render_template('login.html'))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.route('/logout')
@login_required
def logout():
    logout_user()
    # Clear session
    session.clear()
    flash('Anda telah logout!', 'info')
    
    # Buat response dengan header cache-control
    response = make_response(redirect(url_for('login')))
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# ============================================================================
# ADMIN MASTER ROUTES
# ============================================================================
@app.route('/admin-master/dashboard')
@login_required
@admin_master_required
def admin_master_dashboard():
    total_stations = Stasiun.query.count()
    total_opponents = StasiunLawan.query.count()
    
    telkom_opponents = StasiunLawan.query.join(Stasiun).filter(Stasiun.operator == 'telkom').count()
    telkomsel_opponents = StasiunLawan.query.join(Stasiun).filter(Stasiun.operator == 'telkomsel').count()
    indosat_opponents = StasiunLawan.query.join(Stasiun).filter(Stasiun.operator == 'indosat').count()
    xl_opponents = StasiunLawan.query.join(Stasiun).filter(Stasiun.operator == 'xl').count()
    
    telkom_stations = Stasiun.query.filter_by(operator='telkom').count()
    telkomsel_stations = Stasiun.query.filter_by(operator='telkomsel').count()
    indosat_stations = Stasiun.query.filter_by(operator='indosat').count()
    xl_stations = Stasiun.query.filter_by(operator='xl').count()
    
    recent_stations = Stasiun.query.order_by(Stasiun.created_at.desc()).limit(5).all()
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    current_date = datetime.now().strftime('%d %B %Y')
    
    # Get user statistics
    user_stats = {
        'admin_master': User.query.filter_by(role='admin_master').count(),
        'admin_operator': User.query.filter_by(role='admin_operator').count(),
        'user_operator': User.query.filter_by(role='user_operator').count(),
    }
    
    return render_template('admin_master/dashboard.html',
                         total_stations=total_stations,
                         total_opponents=total_opponents,  # WAJIB: Total semua stasiun lawan
                         telkom_stations=telkom_stations,
                         telkomsel_stations=telkomsel_stations,
                         indosat_stations=indosat_stations,
                         xl_stations=xl_stations,
                         # Opsional: kirim juga per operator
                         telkom_opponents=telkom_opponents,
                         telkomsel_opponents=telkomsel_opponents,
                         indosat_opponents=indosat_opponents,
                         xl_opponents=xl_opponents,
                         recent_stations=recent_stations,
                         recent_users=recent_users,
                         user_stats=user_stats,
                         operators=OPERATORS,
                         kota_list=KOTA_LIST,
                         current_date=current_date)

@app.route('/admin-master/users')
@login_required
@admin_master_required
def admin_master_manage_users():
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_master/manage_users.html', users=users)

@app.route('/admin-master/users/add', methods=['GET', 'POST'])
@login_required
@admin_master_required
def admin_master_add_user():
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            role = request.form.get('role')
            operator_type = request.form.get('operator_type')
            
            if not username or not password or not role:
                flash('Username, password, dan role harus diisi!', 'error')
                return redirect(url_for('admin_master_add_user'))
            
            # Cek apakah username sudah ada
            existing_user = User.query.filter_by(username=username).first()
            if existing_user:
                flash('Username sudah digunakan!', 'error')
                return redirect(url_for('admin_master_add_user'))
            
            # Hash password
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            
            # Validasi operator_type berdasarkan role
            if role == 'admin_master':
                operator_type = None
            elif role in ['admin_operator', 'user_operator']:
                if not operator_type or operator_type not in OPERATORS:
                    flash('Operator type harus dipilih!', 'error')
                    return redirect(url_for('admin_master_add_user'))
            
            # Buat user baru
            new_user = User(
                username=username,
                password=hashed_password,
                role=role,
                operator_type=operator_type
            )
            
            db.session.add(new_user)
            db.session.commit()
            
            flash(f'User {username} berhasil ditambahkan!', 'success')
            return redirect(url_for('admin_master_manage_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_master_add_user'))
    
    return render_template('admin_master/add_user.html', 
                         roles=['admin_master', 'admin_operator', 'user_operator'],
                         operators=OPERATORS)

@app.route('/admin-master/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_master_required
def admin_master_edit_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            role = request.form.get('role')
            operator_type = request.form.get('operator_type')
            
            if not username or not role:
                flash('Username dan role harus diisi!', 'error')
                return redirect(url_for('admin_master_edit_user', user_id=user_id))
            
            # Cek apakah username sudah ada (kecuali untuk user ini)
            existing_user = User.query.filter(User.username == username, User.id != user_id).first()
            if existing_user:
                flash('Username sudah digunakan!', 'error')
                return redirect(url_for('admin_master_edit_user', user_id=user_id))
            
            # Update user
            user.username = username
            user.role = role
            
            # Update password jika diisi
            if password:
                user.password = generate_password_hash(password, method='pbkdf2:sha256')
            
            # Update operator_type berdasarkan role
            if role == 'admin_master':
                user.operator_type = None
            elif role in ['admin_operator', 'user_operator']:
                if not operator_type or operator_type not in OPERATORS:
                    flash('Operator type harus dipilih!', 'error')
                    return redirect(url_for('admin_master_edit_user', user_id=user_id))
                user.operator_type = operator_type
            
            db.session.commit()
            
            flash(f'User {username} berhasil diperbarui!', 'success')
            return redirect(url_for('admin_master_manage_users'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_master_edit_user', user_id=user_id))
    
    return render_template('admin_master/edit_user.html', 
                         user=user,
                         roles=['admin_master', 'admin_operator', 'user_operator'],
                         operators=OPERATORS)

@app.route('/admin-master/users/delete/<int:user_id>')
@login_required
@admin_master_required
def admin_master_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    # Tidak boleh menghapus diri sendiri
    if user.id == current_user.id:
        flash('Tidak dapat menghapus akun sendiri!', 'error')
        return redirect(url_for('admin_master_manage_users'))
    
    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.username} berhasil dihapus!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('admin_master_manage_users'))

# ============================================================================
# HAPUS SEMUA DATA (BULK DELETE)
# ============================================================================

@app.route('/admin-master/hapus-semua-data', methods=['POST'])
@login_required
@admin_master_required
def admin_master_hapus_semua_data():
    """Hapus semua data stasiun untuk Admin Master"""
    try:
        # Validasi konfirmasi
        confirm = request.form.get('confirm_delete')
        if confirm != 'HAPUS SEMUA':
            flash('Konfirmasi tidak valid! Ketik "HAPUS SEMUA" untuk melanjutkan.', 'error')
            return redirect(url_for('admin_master_daftar_stasiun'))
        
        # Hitung jumlah data sebelum dihapus
        total_stations = Stasiun.query.count()
        total_opponents = StasiunLawan.query.count()
        total_uploads = UploadGambar.query.count()
        total_status = StatusUpdate.query.count()
        total_groups = GrupStasiun.query.count()
        
        print(f"🗑️ MEMULAI PROSES HAPUS SEMUA DATA - ADMIN MASTER")
        print(f"   - Stasiun: {total_stations}")
        print(f"   - Stasiun Lawan: {total_opponents}")
        print(f"   - Upload Gambar: {total_uploads}")
        print(f"   - Status Update: {total_status}")
        print(f"   - Grup Stasiun: {total_groups}")
        
        # ===== 1. HAPUS SEMUA FILE DARI CLOUDINARY =====
        all_uploads = UploadGambar.query.all()
        cloudinary_deleted = 0
        cloudinary_failed = 0
        
        # Koleksi public_id unik
        unique_public_ids = set()
        for upload in all_uploads:
            unique_public_ids.add(upload.public_id)
        
        print(f"   - File unik di Cloudinary: {len(unique_public_ids)}")
        
        for public_id in unique_public_ids:
            try:
                success, msg = delete_from_cloudinary(public_id)
                if success:
                    cloudinary_deleted += 1
                    print(f"     ✓ {public_id}")
                else:
                    cloudinary_failed += 1
                    print(f"     ✗ {public_id}: {msg}")
            except Exception as e:
                cloudinary_failed += 1
                print(f"     ✗ {public_id}: {e}")
        
        # ===== 2. HAPUS SEMUA DATA DARI DATABASE =====
        # Hapus status updates
        StatusUpdate.query.delete()
        
        # Hapus upload gambar
        UploadGambar.query.delete()
        
        # Hapus stasiun lawan
        StasiunLawan.query.delete()
        
        # Hapus grup stasiun
        GrupStasiun.query.delete()
        
        # Hapus stasiun
        Stasiun.query.delete()
        
        # Commit perubahan
        db.session.commit()
        
        # ===== 3. TAMPILKAN HASIL =====
        message = f"""
        ✅ BERHASIL MENGHAPUS SEMUA DATA!
        
        📊 RINCIAN PENGHAPUSAN:
        • Stasiun: {total_stations} data
        • Stasiun Lawan: {total_opponents} data
        • Upload Gambar: {total_uploads} data
        • Status Update: {total_status} data
        • Grup Stasiun: {total_groups} data
        
        ☁️ CLOUDINARY:
        • File berhasil dihapus: {cloudinary_deleted} file
        • File gagal dihapus: {cloudinary_failed} file
        """
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        error_msg = f"❌ ERROR: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        flash(error_msg, 'error')
    
    return redirect(url_for('admin_master_daftar_stasiun'))

@app.route('/admin-operator/hapus-semua-data', methods=['POST'])
@login_required
@admin_operator_required
def admin_operator_hapus_semua_data():
    """Hapus semua data stasiun untuk operator tertentu"""
    try:
        operator = current_user.operator_type
        
        # Validasi konfirmasi
        confirm = request.form.get('confirm_delete')
        expected_confirm = f'HAPUS {operator.upper()}'
        if confirm != expected_confirm:
            flash(f'Konfirmasi tidak valid! Ketik "{expected_confirm}" untuk melanjutkan.', 'error')
            return redirect(url_for('admin_operator_daftar_stasiun'))
        
        # Hitung jumlah data sebelum dihapus
        stations = Stasiun.query.filter_by(operator=operator).all()
        station_ids = [s.id for s in stations]
        
        total_stations = len(stations)
        total_opponents = StasiunLawan.query.filter(StasiunLawan.stasiun_id.in_(station_ids)).count() if station_ids else 0
        
        # Hitung upload dan status
        total_uploads = 0
        total_status = 0
        total_groups = 0
        
        if station_ids:
            total_uploads = UploadGambar.query.filter(UploadGambar.stasiun_id.in_(station_ids)).count()
            total_groups = GrupStasiun.query.filter(GrupStasiun.stasiun_id.in_(station_ids)).count()
            
            # Hitung status updates
            opponent_ids = [o.id for o in StasiunLawan.query.filter(StasiunLawan.stasiun_id.in_(station_ids)).all()]
            if opponent_ids:
                total_status = StatusUpdate.query.filter(StatusUpdate.stasiun_lawan_id.in_(opponent_ids)).count()
        
        print(f"🗑️ MEMULAI PROSES HAPUS SEMUA DATA - OPERATOR {operator.upper()}")
        print(f"   - Stasiun: {total_stations}")
        print(f"   - Stasiun Lawan: {total_opponents}")
        print(f"   - Upload Gambar: {total_uploads}")
        print(f"   - Status Update: {total_status}")
        print(f"   - Grup Stasiun: {total_groups}")
        
        if not station_ids:
            flash(f'Tidak ada data untuk operator {operator.upper()}', 'info')
            return redirect(url_for('admin_operator_daftar_stasiun'))
        
        # ===== 1. HAPUS SEMUA FILE DARI CLOUDINARY =====
        all_uploads = UploadGambar.query.filter(UploadGambar.stasiun_id.in_(station_ids)).all()
        cloudinary_deleted = 0
        cloudinary_failed = 0
        
        # Koleksi public_id unik
        unique_public_ids = set()
        for upload in all_uploads:
            unique_public_ids.add(upload.public_id)
        
        print(f"   - File unik di Cloudinary: {len(unique_public_ids)}")
        
        for public_id in unique_public_ids:
            try:
                # Cek apakah public_id ini masih digunakan oleh operator lain
                other_refs = UploadGambar.query\
                    .filter(UploadGambar.public_id == public_id,
                           ~UploadGambar.stasiun_id.in_(station_ids))\
                    .count()
                
                if other_refs == 0:
                    success, msg = delete_from_cloudinary(public_id)
                    if success:
                        cloudinary_deleted += 1
                        print(f"     ✓ {public_id}")
                    else:
                        cloudinary_failed += 1
                        print(f"     ✗ {public_id}: {msg}")
                else:
                    print(f"     ⚠ {public_id} masih digunakan operator lain, dilewati")
                    cloudinary_deleted += 1  # Hitung sebagai berhasil karena tidak perlu dihapus
            except Exception as e:
                cloudinary_failed += 1
                print(f"     ✗ {public_id}: {e}")
        
        # ===== 2. HAPUS DATA DARI DATABASE =====
        # Hapus status updates
        if opponent_ids:
            StatusUpdate.query.filter(StatusUpdate.stasiun_lawan_id.in_(opponent_ids)).delete()
        
        # Hapus upload gambar
        UploadGambar.query.filter(UploadGambar.stasiun_id.in_(station_ids)).delete()
        
        # Hapus stasiun lawan
        StasiunLawan.query.filter(StasiunLawan.stasiun_id.in_(station_ids)).delete()
        
        # Hapus grup stasiun
        GrupStasiun.query.filter(GrupStasiun.stasiun_id.in_(station_ids)).delete()
        
        # Hapus stasiun
        for station in stations:
            db.session.delete(station)
        
        # Commit perubahan
        db.session.commit()
        
        # ===== 3. TAMPILKAN HASIL =====
        message = f"""
        ✅ BERHASIL MENGHAPUS SEMUA DATA OPERATOR {operator.upper()}!
        
        📊 RINCIAN PENGHAPUSAN:
        • Stasiun: {total_stations} data
        • Stasiun Lawan: {total_opponents} data
        • Upload Gambar: {total_uploads} data
        • Status Update: {total_status} data
        • Grup Stasiun: {total_groups} data
        
        ☁️ CLOUDINARY:
        • File berhasil dihapus: {cloudinary_deleted} file
        • File gagal dihapus: {cloudinary_failed} file
        """
        
        flash(message, 'success')
        
    except Exception as e:
        db.session.rollback()
        error_msg = f"❌ ERROR: {str(e)}"
        print(error_msg)
        import traceback
        traceback.print_exc()
        flash(error_msg, 'error')
    
    return redirect(url_for('admin_operator_daftar_stasiun'))

@app.route('/admin-master/tambah-data', methods=['GET', 'POST'])
@login_required
@admin_master_required
def admin_master_tambah_data():
    if request.method == 'POST':
        try:
            stasiun_name = request.form.get('stasiun_name').strip()
            operator = request.form.get('operator')
            kota = request.form.get('kota')
            
            stasiun_lawan_list = request.form.getlist('stasiun_lawan[]')
            freq_tx_list = request.form.getlist('freq_tx[]')
            freq_rx_list = request.form.getlist('freq_rx[]')
            
            if not stasiun_name:
                flash('Nama stasiun harus diisi!', 'error')
                return redirect(url_for('admin_master_tambah_data'))
            
            new_station = Stasiun(
                stasiun_name=stasiun_name,
                operator=operator,
                kota=kota,
                created_by=current_user.id
            )
            db.session.add(new_station)
            db.session.flush()
            
            for i, nama_lawan in enumerate(stasiun_lawan_list):
                nama_lawan = nama_lawan.strip()
                if nama_lawan:
                    freq_tx = freq_tx_list[i].strip() if i < len(freq_tx_list) else ''
                    freq_rx = freq_rx_list[i].strip() if i < len(freq_rx_list) else ''
                    
                    new_lawan = StasiunLawan(
                        stasiun_id=new_station.id,
                        nama_stasiun_lawan=nama_lawan,
                        freq_tx=freq_tx if freq_tx else None,
                        freq_rx=freq_rx if freq_rx else None,
                        group_id=None,
                        urutan=i
                    )
                    db.session.add(new_lawan)
            
            db.session.commit()
            flash(f'Stasiun "{stasiun_name}" berhasil ditambahkan!', 'success')
            return redirect(url_for('admin_master_daftar_stasiun'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_master_tambah_data'))
    
    return render_template('admin_master/tambah_data.html',
                         operators=OPERATORS,
                         kota_list=KOTA_LIST)

@app.route('/admin-master/daftar-stasiun')
@login_required
@admin_master_required
def admin_master_daftar_stasiun():
    operator = request.args.get('operator', 'all')
    kota = request.args.get('kota', 'all')
    search_stasiun = request.args.get('search_stasiun', '').strip()
    
    query = Stasiun.query
    
    if operator != 'all':
        query = query.filter_by(operator=operator)
    
    if kota != 'all':
        query = query.filter_by(kota=kota)
    
    if search_stasiun:
        query = query.filter(Stasiun.stasiun_name.ilike(f'%{search_stasiun}%'))
    
    per_page = request.args.get('per_page', 10, type=int)
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    page = request.args.get('page', 1, type=int)
    stations_pagination = query.order_by(Stasiun.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    stations = stations_pagination.items
    for station in stations:
        station.stasiun_lawan_list = StasiunLawan.query\
            .filter_by(stasiun_id=station.id)\
            .order_by(StasiunLawan.urutan)\
            .all()
        
        for lawan in station.stasiun_lawan_list:
            lawan.latest_status = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=lawan.id)\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
        
        station.uploads = UploadGambar.query\
            .filter_by(stasiun_id=station.id)\
            .all()
    
    return render_template('admin_master/daftar_stasiun.html',
                         stations=stations,
                         pagination=stations_pagination,
                         operators=OPERATORS,
                         kota_list=KOTA_LIST,
                         selected_operator=operator,
                         selected_kota=kota,
                         status_options=STATUS_OPTIONS)

@app.route('/admin-master/stasiun/<int:station_id>')
@login_required
@admin_master_required
def admin_master_detail_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    # Ambil semua lawan
    all_lawans = StasiunLawan.query\
        .filter_by(stasiun_id=station_id)\
        .order_by(StasiunLawan.urutan)\
        .all()
    
    # Ambil semua grup dari tabel GrupStasiun
    all_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
    groups_dict = {g.id: g.nama_grup for g in all_groups}
    
    # Ambil semua upload UNIK (group by public_id dan group_id)
    uploads_query = db.session.query(
        UploadGambar.group_id,
        UploadGambar.public_id,
        db.func.max(UploadGambar.id).label('latest_id')
    ).filter_by(stasiun_id=station_id)\
     .group_by(UploadGambar.group_id, UploadGambar.public_id)\
     .subquery()
    
    all_uploads = UploadGambar.query\
        .join(uploads_query, UploadGambar.id == uploads_query.c.latest_id)\
        .order_by(UploadGambar.uploaded_at.desc())\
        .all()
    
    # Kelompokkan lawan dan upload berdasarkan grup
    groups = {}
    
    # ===== PROSES LAWAN PER GRUP =====
    for lawan in all_lawans:
        group_key = lawan.group_id if lawan.group_id is not None else 'ungrouped'
        
        if group_key not in groups:
            groups[group_key] = {
                'group_id': lawan.group_id,
                'nama_grup': groups_dict.get(lawan.group_id) if lawan.group_id else None,
                'opponents': [],
                'uploads': [],  # PASTIKAN FIELD INI ADA
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        # Ambil status terbaru
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=lawan.id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        lawan_data = {
            'id': lawan.id,
            'nama': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'latest_status': latest_status.status if latest_status else None,
            'catatan': latest_status.catatan if latest_status else '',
            'status_updated_at': latest_status.updated_at if latest_status else None
        }
        
        groups[group_key]['opponents'].append(lawan_data)
        
        # Hitung statistik
        if latest_status and latest_status.status in groups[group_key]['status_summary']:
            groups[group_key]['status_summary'][latest_status.status] += 1
    
    # ===== PROSES UPLOAD PER GRUP - YANG INI PENTING! =====
    for upload in all_uploads:
        group_key = upload.group_id if upload.group_id is not None else 'ungrouped'
        
        if group_key not in groups:
            groups[group_key] = {
                'group_id': upload.group_id,
                'nama_grup': groups_dict.get(upload.group_id) if upload.group_id else None,
                'opponents': [],
                'uploads': [],
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        # Cek duplikasi public_id dalam grup yang sama
        existing_public_ids = [u['public_id'] for u in groups[group_key]['uploads']]
        if upload.public_id not in existing_public_ids:
            upload_data = {
                'id': upload.id,
                'public_id': upload.public_id,
                'cloudinary_url': upload.cloudinary_url,
                'original_filename': upload.original_filename,
                'uploaded_at': upload.uploaded_at,
                'format': upload.format,
                'width': upload.width,
                'height': upload.height,
                'bytes_size': upload.bytes_size
            }
            groups[group_key]['uploads'].append(upload_data)
    
    # Urutkan grup: ungrouped dulu, lalu grup numeric
    sorted_groups = {}
    
    # Ungrouped dulu
    if 'ungrouped' in groups:
        sorted_groups['ungrouped'] = groups.pop('ungrouped')
    
    # Grup numeric diurutkan
    numeric_groups = {}
    for key, data in groups.items():
        if key != 'ungrouped':
            try:
                numeric_groups[int(key)] = data
            except:
                numeric_groups[key] = data
    
    for key in sorted(numeric_groups.keys()):
        sorted_groups[str(key)] = numeric_groups[key]
    
    # Hitung total opponents
    total_opponents = len(all_lawans)
    
    # Tambahkan station.grup_list untuk digunakan di template
    station.grup_list = all_groups
    
    return render_template('admin_master/detail_stasiun.html',
                         station=station,
                         groups=sorted_groups,
                         all_uploads=all_uploads,
                         total_opponents=total_opponents)

@app.route('/admin-master/edit-stasiun/<int:station_id>', methods=['GET', 'POST'])
@login_required
@admin_master_required
def admin_master_edit_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if request.method == 'POST':
        try:
            # Validasi input
            stasiun_name = request.form.get('stasiun_name', '').strip()
            operator = request.form.get('operator', '').strip()
            kota = request.form.get('kota', '').strip()
            
            if not stasiun_name:
                flash('❌ Nama stasiun harus diisi!', 'error')
                return redirect(url_for('admin_master_edit_stasiun', station_id=station_id))
            
            if not operator:
                flash('❌ Operator harus dipilih!', 'error')
                return redirect(url_for('admin_master_edit_stasiun', station_id=station_id))
            
            if not kota:
                flash('❌ Kota/Kabupaten harus dipilih!', 'error')
                return redirect(url_for('admin_master_edit_stasiun', station_id=station_id))
            
            # Catat perubahan untuk pesan
            changes = []
            
            # 1. Update data stasiun utama
            if station.stasiun_name != stasiun_name:
                changes.append(f"📝 Nama stasiun: '{station.stasiun_name}' → '{stasiun_name}'")
                station.stasiun_name = stasiun_name
            
            if station.operator != operator:
                changes.append(f"🔄 Operator: '{station.operator}' → '{operator}'")
                station.operator = operator
            
            if station.kota != kota:
                changes.append(f"📍 Kota: '{station.kota}' → '{kota}'")
                station.kota = kota
            
            # 2. Ambil data dari form
            lawan_ids = request.form.getlist('lawan_id[]')
            stasiun_lawan_list = request.form.getlist('stasiun_lawan[]')
            freq_tx_list = request.form.getlist('freq_tx[]')
            freq_rx_list = request.form.getlist('freq_rx[]')
            deleted_lawan_ids = request.form.getlist('deleted_lawan_ids[]')
            
            # Validasi minimal 1 stasiun lawan
            valid_lawan = [n for n in stasiun_lawan_list if n.strip()]
            if not valid_lawan:
                flash('❌ Minimal harus ada 1 stasiun lawan!', 'error')
                return redirect(url_for('admin_master_edit_stasiun', station_id=station_id))
            
            # 3. Hapus lawan yang ditandai untuk dihapus
            deleted_count = 0
            if deleted_lawan_ids:
                for lawan_id in deleted_lawan_ids:
                    if lawan_id:
                        lawan = StasiunLawan.query.get(int(lawan_id))
                        if lawan and lawan.stasiun_id == station_id:
                            
                            # Putuskan relasi upload, jangan hapus file
                            uploads = UploadGambar.query.filter_by(stasiun_lawan_id=lawan.id).all()
                            for upload in uploads:
                                upload.stasiun_lawan_id = None
                            
                            # Hapus status updates
                            StatusUpdate.query.filter_by(stasiun_lawan_id=lawan.id).delete()
                            
                            changes.append(f"❌ HAPUS: '{lawan.nama_stasiun_lawan}' (status dihapus, gambar tetap ada)")
                            
                            db.session.delete(lawan)
                            deleted_count += 1
            
            # 4. Proses data yang ada (UPDATE) dan data baru (TAMBAH)
            updated_count = 0
            new_count = 0
            freq_tx_updated_count = 0
            freq_rx_updated_count = 0
            
            for i, nama_lawan in enumerate(stasiun_lawan_list):
                nama_lawan = nama_lawan.strip()
                if not nama_lawan:
                    continue
                
                freq_tx = freq_tx_list[i].strip() if i < len(freq_tx_list) else ''
                freq_rx = freq_rx_list[i].strip() if i < len(freq_rx_list) else ''
                
                # Cek apakah ini data existing (punya ID) atau baru
                if i < len(lawan_ids) and lawan_ids[i]:
                    # UPDATE DATA EXISTING
                    lawan_id = int(lawan_ids[i])
                    lawan = StasiunLawan.query.get(lawan_id)
                    
                    if lawan and lawan.stasiun_id == station_id:
                        # Cek perubahan nama
                        if lawan.nama_stasiun_lawan != nama_lawan:
                            changes.append(f"✏️ Ubah nama: '{lawan.nama_stasiun_lawan}' → '{nama_lawan}'")
                            lawan.nama_stasiun_lawan = nama_lawan
                            updated_count += 1
                        
                        # Cek perubahan frekuensi TX
                        old_freq_tx = lawan.freq_tx or ''
                        new_freq_tx = freq_tx if freq_tx else ''
                        if old_freq_tx != new_freq_tx:
                            if old_freq_tx and new_freq_tx:
                                changes.append(f"📡 Ubah Freq Tx '{lawan.nama_stasiun_lawan}': '{old_freq_tx}' → '{new_freq_tx}'")
                            elif new_freq_tx and not old_freq_tx:
                                changes.append(f"📡 Tambah Freq Tx '{lawan.nama_stasiun_lawan}': '{new_freq_tx}'")
                            elif old_freq_tx and not new_freq_tx:
                                changes.append(f"📡 Hapus Freq Tx '{lawan.nama_stasiun_lawan}' (sebelumnya '{old_freq_tx}')")
                            
                            lawan.freq_tx = new_freq_tx if new_freq_tx else None
                            freq_tx_updated_count += 1
                        
                        # Cek perubahan frekuensi RX
                        old_freq_rx = lawan.freq_rx or ''
                        new_freq_rx = freq_rx if freq_rx else ''
                        if old_freq_rx != new_freq_rx:
                            if old_freq_rx and new_freq_rx:
                                changes.append(f"📡 Ubah Freq Rx '{lawan.nama_stasiun_lawan}': '{old_freq_rx}' → '{new_freq_rx}'")
                            elif new_freq_rx and not old_freq_rx:
                                changes.append(f"📡 Tambah Freq Rx '{lawan.nama_stasiun_lawan}': '{new_freq_rx}'")
                            elif old_freq_rx and not new_freq_rx:
                                changes.append(f"📡 Hapus Freq Rx '{lawan.nama_stasiun_lawan}' (sebelumnya '{old_freq_rx}')")
                            
                            lawan.freq_rx = new_freq_rx if new_freq_rx else None
                            freq_rx_updated_count += 1
                else:
                    # DATA BARU
                    last_urutan = db.session.query(db.func.max(StasiunLawan.urutan))\
                        .filter_by(stasiun_id=station_id).scalar() or 0
                    
                    new_lawan = StasiunLawan(
                        stasiun_id=station_id,
                        nama_stasiun_lawan=nama_lawan,
                        freq_tx=freq_tx if freq_tx else None,
                        freq_rx=freq_rx if freq_rx else None,
                        group_id=None,
                        urutan=last_urutan + 1
                    )
                    db.session.add(new_lawan)
                    new_count += 1
                    freq_info = []
                    if freq_tx:
                        freq_info.append(f"Tx:{freq_tx}")
                    if freq_rx:
                        freq_info.append(f"Rx:{freq_rx}")
                    freq_str = f" ({', '.join(freq_info)})" if freq_info else ""
                    changes.append(f"✅ TAMBAH: '{nama_lawan}'{freq_str}")
            
            # 5. Commit semua perubahan
            db.session.commit()
            
            # 6. Tampilkan ringkasan perubahan
            if changes:
                flash('📋 RINCIAN PERUBAHAN:', 'info')
                for change in changes:
                    flash(change, 'debug')
            
            # Ringkasan statistik
            summary = []
            if new_count > 0:
                summary.append(f"✅ {new_count} lawan baru")
            if updated_count > 0:
                summary.append(f"✏️ {updated_count} nama diubah")
            if freq_tx_updated_count > 0:
                summary.append(f"📡 {freq_tx_updated_count} Freq Tx diubah")
            if freq_rx_updated_count > 0:
                summary.append(f"📡 {freq_rx_updated_count} Freq Rx diubah")
            if deleted_count > 0:
                summary.append(f"❌ {deleted_count} lawan dihapus")
            
            if summary:
                flash(' ✓ '.join(summary), 'success')
            else:
                flash('ℹ️ Tidak ada perubahan data.', 'info')
            
            return redirect(url_for('admin_master_detail_stasiun', station_id=station_id))
            
        except Exception as e:
            db.session.rollback()
            error_msg = str(e)
            flash(f'❌ ERROR: {error_msg}', 'error')
            print(f"ERROR in admin_master_edit_stasiun: {error_msg}")
            import traceback
            traceback.print_exc()
    
    return render_template('admin_master/edit_stasiun.html',
                         station=station,
                         operators=OPERATORS,
                         kota_list=KOTA_LIST)

@app.route('/admin-master/hapus-stasiun/<int:station_id>')
@login_required
@admin_master_required
def admin_master_hapus_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    try:
        # 1. HAPUS DULU SEMUA GRUP YANG TERKAIT DENGAN STASIUN INI
        groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
        for group in groups:
            # Update lawan yang merujuk ke grup ini
            StasiunLawan.query.filter_by(grup_id=group.id).update({
                'group_id': None,
                'grup_id': None
            })
            
            # Update upload yang merujuk ke grup ini
            UploadGambar.query.filter_by(group_id=group.id).update({
                'group_id': None
            })
            
            # Hapus grup
            db.session.delete(group)
        
        # 2. AMBIL SEMUA LAWAN
        opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).all()
        opponent_ids = [opponent.id for opponent in opponents]
        
        # 3. HAPUS STATUS UPDATES
        if opponent_ids:
            StatusUpdate.query.filter(StatusUpdate.stasiun_lawan_id.in_(opponent_ids)).delete()
        
        # 4. HAPUS UPLOAD GAMBAR (FILE FISIK DAN RECORD) - UPDATE UNTUK CLOUDINARY
        all_uploads = UploadGambar.query.filter_by(stasiun_id=station_id).all()
        for upload in all_uploads:
            # Hapus dari Cloudinary
            delete_from_cloudinary(upload.public_id)
            db.session.delete(upload)
        
        # 5. HAPUS STASIUN LAWAN
        StasiunLawan.query.filter_by(stasiun_id=station_id).delete()
        
        # 6. TERAKHIR, HAPUS STASIUN
        db.session.delete(station)
        db.session.commit()
        
        flash(f'Stasiun "{station.stasiun_name}" berhasil dihapus!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
        print(f"Error menghapus stasiun: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('admin_master_daftar_stasiun'))

@app.route('/admin-master/upload-excel', methods=['GET', 'POST'])
@login_required
@admin_master_required
def admin_master_upload_excel():
    """Halaman upload Excel untuk admin master"""
    
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.files:
                flash('Tidak ada file yang dipilih!', 'error')
                return redirect(url_for('admin_master_upload_excel'))
            
            file = request.files['excel_file']
            
            if file.filename == '':
                flash('Tidak ada file yang dipilih!', 'error')
                return redirect(url_for('admin_master_upload_excel'))
            
            # Validasi file
            is_valid, error_msg = validate_excel_file(file)
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(url_for('admin_master_upload_excel'))
            
            # Get kota default
            kota_default = request.form.get('kota_default', 'samarinda')
            
            # Simpan file sementara
            temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_filename = f"admin_master_{timestamp}_{secure_filename(file.filename)}"
            temp_path = os.path.join(temp_dir, temp_filename)
            
            file.save(temp_path)
            
            # Process file
            success_count, error_count, messages = process_excel_upload_admin_master(
                temp_path, 
                current_user,
                kota_default
            )
            
            # Hapus file temp
            try:
                os.remove(temp_path)
            except:
                pass
            
            # Tampilkan hasil
            for msg in messages:
                if 'Error' in msg or 'error' in msg:
                    flash(msg, 'error')
                elif 'sukses' in msg or 'berhasil' in msg:
                    flash(msg, 'success')
                else:
                    flash(msg, 'info')
            
            return redirect(url_for('admin_master_daftar_stasiun'))
            
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_master_upload_excel'))
    
    # GET request
    total_stations = Stasiun.query.count()
    
    # Hitung total operator unik
    operators_count = db.session.query(Stasiun.operator).distinct().count()
    
    return render_template('admin_master/upload_excel.html',
                         total_stations=total_stations,
                         total_operators=operators_count,
                         operators=OPERATORS,
                         kota_list=KOTA_LIST)

@app.route('/admin-master/download-template')
@login_required
@admin_master_required
def admin_master_download_template():
    """Download template Excel untuk admin master"""
    return generate_admin_master_template()

@app.route('/debug-check-duplicates')
@login_required
@admin_master_required
def debug_check_duplicates():
    """Cek apakah ada file dengan nama yang sama - UPDATE untuk Cloudinary"""
    uploads = UploadGambar.query.all()
    
    # Kelompokkan berdasarkan original_filename
    filename_groups = {}
    for upload in uploads:
        original_name = upload.original_filename or 'unknown'
        
        if original_name not in filename_groups:
            filename_groups[original_name] = []
        filename_groups[original_name].append(upload)
    
    result = []
    result.append("=== CEK DUPLIKASI NAMA FILE ===")
    result.append("")
    
    for original_name, uploads_list in filename_groups.items():
        if len(uploads_list) > 1:
            result.append(f"⚠️ {original_name} - {len(uploads_list)} file:")
            for upload in uploads_list:
                station = Stasiun.query.get(upload.stasiun_id)
                result.append(f"   - {upload.public_id} (Stasiun: {station.stasiun_name if station else 'Unknown'})")
            result.append("")
    
    return "<pre>" + "\n".join(result) + "</pre>"

# ============================================================================
# ADMIN OPERATOR ROUTES
# ============================================================================

@app.route('/admin-operator/dashboard')
@login_required
@admin_operator_required
def admin_operator_dashboard():
    operator = current_user.operator_type
    
    total_stations = Stasiun.query.filter_by(operator=operator).count()
    
    # PERBAIKAN: Hitung SEMUA stasiun lawan (sama seperti user operator)
    total_opponents = StasiunLawan.query\
        .join(Stasiun)\
        .filter(Stasiun.operator == operator)\
        .count()
    
    # Hitung stasiun yang sudah upload
    stations_with_uploads = Stasiun.query\
        .filter(Stasiun.operator == operator, Stasiun.uploads.any())\
        .count()
    
    recent_stations = Stasiun.query\
        .filter_by(operator=operator)\
        .order_by(Stasiun.created_at.desc())\
        .limit(5).all()
    
    kota_for_operator = db.session.query(Stasiun.kota)\
        .filter_by(operator=operator)\
        .distinct()\
        .all()
    kota_list_operator = [k[0] for k in kota_for_operator]
    
    current_date = datetime.now().strftime('%d %B %Y')
    
    return render_template('admin_operator/dashboard.html',
                         operator=operator,
                         operator_name=operator.upper(),
                         current_date=current_date,
                         total_stations=total_stations,
                         total_opponents=total_opponents,  # SEKARANG SAMA
                         stations_with_uploads=stations_with_uploads,
                         recent_stations=recent_stations,
                         kota_list=kota_list_operator)

@app.route('/admin-operator/tambah-data', methods=['GET', 'POST'])
@login_required
@admin_operator_required
def admin_operator_tambah_data():
    operator = current_user.operator_type
    
    if request.method == 'POST':
        try:
            stasiun_name = request.form.get('stasiun_name').strip()
            kota = request.form.get('kota')
            
            stasiun_lawan_list = request.form.getlist('stasiun_lawan[]')
            freq_tx_list = request.form.getlist('freq_tx[]')
            freq_rx_list = request.form.getlist('freq_rx[]')
            
            if not stasiun_name:
                flash('Nama stasiun harus diisi!', 'error')
                return redirect(url_for('admin_operator_tambah_data'))
            
            # Validasi minimal 1 stasiun lawan
            valid_lawan = [n for n in stasiun_lawan_list if n.strip()]
            if not valid_lawan:
                flash('Minimal harus ada 1 stasiun lawan!', 'error')
                return redirect(url_for('admin_operator_tambah_data'))
            
            new_station = Stasiun(
                stasiun_name=stasiun_name,
                operator=operator,
                kota=kota,
                created_by=current_user.id
            )
            db.session.add(new_station)
            db.session.flush()
            
            for i, nama_lawan in enumerate(stasiun_lawan_list):
                nama_lawan = nama_lawan.strip()
                if nama_lawan:
                    freq_tx = freq_tx_list[i].strip() if i < len(freq_tx_list) else ''
                    freq_rx = freq_rx_list[i].strip() if i < len(freq_rx_list) else ''
                    
                    new_lawan = StasiunLawan(
                        stasiun_id=new_station.id,
                        nama_stasiun_lawan=nama_lawan,
                        freq_tx=freq_tx if freq_tx else None,
                        freq_rx=freq_rx if freq_rx else None,
                        group_id=None,
                        urutan=i
                    )
                    db.session.add(new_lawan)
            
            db.session.commit()
            flash(f'Stasiun "{stasiun_name}" berhasil ditambahkan!', 'success')
            return redirect(url_for('admin_operator_daftar_stasiun'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_operator_tambah_data'))
    
    return render_template('admin_operator/tambah_data.html',
                         kota_list=KOTA_LIST,
                         operator=operator)

@app.route('/admin-operator/daftar-stasiun')
@login_required
@admin_operator_required
def admin_operator_daftar_stasiun():
    operator = current_user.operator_type
    kota = request.args.get('kota', 'all')
    search_stasiun = request.args.get('search_stasiun', '').strip()
    
    query = Stasiun.query.filter_by(operator=operator)
    
    if kota != 'all':
        query = query.filter_by(kota=kota)
    
    if search_stasiun:
        query = query.filter(Stasiun.stasiun_name.ilike(f'%{search_stasiun}%'))
    
    per_page = request.args.get('per_page', 10, type=int)
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    page = request.args.get('page', 1, type=int)
    stations_pagination = query.order_by(Stasiun.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    stations = stations_pagination.items
    for station in stations:
        station.stasiun_lawan_list = StasiunLawan.query\
            .filter_by(stasiun_id=station.id)\
            .order_by(StasiunLawan.urutan)\
            .all()
        
        station.grup_list = GrupStasiun.query.filter_by(stasiun_id=station.id).all()
        
        for lawan in station.stasiun_lawan_list:
            lawan.latest_status = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=lawan.id)\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
        
        station.uploads = UploadGambar.query\
            .filter_by(stasiun_id=station.id)\
            .all()
    
    kota_for_operator = db.session.query(Stasiun.kota)\
        .filter_by(operator=operator)\
        .distinct()\
        .all()
    kota_list_operator = [k[0] for k in kota_for_operator]
    
    return render_template('admin_operator/daftar_stasiun.html',
                         stations=stations,
                         pagination=stations_pagination,
                         operator=operator,
                         operator_name=operator.upper(),
                         kota_list=kota_list_operator,
                         selected_kota=kota,
                         status_options=STATUS_OPTIONS)

@app.route('/admin-operator/stasiun/<int:station_id>')
@login_required
@admin_operator_required
def admin_operator_detail_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Stasiun ini bukan milik operator Anda!', 'error')
        return redirect(url_for('admin_operator_dashboard'))
    
    # Ambil semua lawan
    all_lawans = StasiunLawan.query\
        .filter_by(stasiun_id=station_id)\
        .order_by(StasiunLawan.urutan)\
        .all()
    
    # Ambil semua grup dari tabel GrupStasiun
    all_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
    groups_dict = {g.id: g.nama_grup for g in all_groups}
    
    # Ambil semua upload UNIK (group by public_id dan group_id)
    uploads_query = db.session.query(
        UploadGambar.group_id,
        UploadGambar.public_id,
        db.func.max(UploadGambar.id).label('latest_id')
    ).filter_by(stasiun_id=station_id)\
     .group_by(UploadGambar.group_id, UploadGambar.public_id)\
     .subquery()
    
    all_uploads = UploadGambar.query\
        .join(uploads_query, UploadGambar.id == uploads_query.c.latest_id)\
        .order_by(UploadGambar.uploaded_at.desc())\
        .all()
    
    # Kelompokkan lawan dan upload berdasarkan grup
    groups = {}
    
    # ===== PROSES LAWAN PER GRUP =====
    for lawan in all_lawans:
        group_key = lawan.group_id if lawan.group_id is not None else 'ungrouped'
        
        if group_key not in groups:
            groups[group_key] = {
                'group_id': lawan.group_id,
                'nama_grup': groups_dict.get(lawan.group_id) if lawan.group_id else None,
                'opponents': [],
                'uploads': [],  # PASTIKAN FIELD INI ADA
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        # Ambil status terbaru
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=lawan.id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        lawan_data = {
            'id': lawan.id,
            'nama': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'latest_status': latest_status.status if latest_status else None,
            'catatan': latest_status.catatan if latest_status else '',
            'status_updated_at': latest_status.updated_at if latest_status else None
        }
        
        groups[group_key]['opponents'].append(lawan_data)
        
        # Hitung statistik
        if latest_status and latest_status.status in groups[group_key]['status_summary']:
            groups[group_key]['status_summary'][latest_status.status] += 1
    
    # ===== PROSES UPLOAD PER GRUP - INI PENTING! =====
    for upload in all_uploads:
        group_key = upload.group_id if upload.group_id is not None else 'ungrouped'
        
        if group_key not in groups:
            groups[group_key] = {
                'group_id': upload.group_id,
                'nama_grup': groups_dict.get(upload.group_id) if upload.group_id else None,
                'opponents': [],
                'uploads': [],
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        # Cek duplikasi public_id dalam grup yang sama
        existing_public_ids = [u['public_id'] for u in groups[group_key]['uploads']]
        if upload.public_id not in existing_public_ids:
            upload_data = {
                'id': upload.id,
                'public_id': upload.public_id,
                'cloudinary_url': upload.cloudinary_url,
                'original_filename': upload.original_filename,
                'uploaded_at': upload.uploaded_at,
                'format': upload.format,
                'width': upload.width,
                'height': upload.height,
                'bytes_size': upload.bytes_size
            }
            groups[group_key]['uploads'].append(upload_data)
    
    # Urutkan grup: ungrouped dulu, lalu grup numeric
    sorted_groups = {}
    
    # Ungrouped dulu
    if 'ungrouped' in groups:
        sorted_groups['ungrouped'] = groups.pop('ungrouped')
    
    # Grup numeric diurutkan
    numeric_groups = {}
    for key, data in groups.items():
        if key != 'ungrouped':
            try:
                numeric_groups[int(key)] = data
            except:
                numeric_groups[key] = data
    
    for key in sorted(numeric_groups.keys()):
        sorted_groups[str(key)] = numeric_groups[key]
    
    # Tambahkan station.grup_list untuk digunakan di template
    station.grup_list = all_groups
    
    return render_template('admin_operator/detail_stasiun.html',
                         station=station,
                         groups=sorted_groups,
                         all_uploads=all_uploads,
                         total_opponents=len(all_lawans))

@app.route('/admin-operator/edit-stasiun/<int:station_id>', methods=['GET', 'POST'])
@login_required
@admin_operator_required
def admin_operator_edit_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('❌ Akses ditolak! Stasiun ini bukan milik operator Anda.', 'error')
        return redirect(url_for('admin_operator_dashboard'))
    
    if request.method == 'POST':
        try:
            # Validasi input
            stasiun_name = request.form.get('stasiun_name', '').strip()
            kota = request.form.get('kota', '').strip()
            
            if not stasiun_name:
                flash('❌ Nama stasiun harus diisi!', 'error')
                return redirect(url_for('admin_operator_edit_stasiun', station_id=station_id))
            
            if not kota:
                flash('❌ Kota/Kabupaten harus dipilih!', 'error')
                return redirect(url_for('admin_operator_edit_stasiun', station_id=station_id))
            
            # Catat perubahan untuk pesan
            changes = []
            
            # 1. Update data stasiun utama
            if station.stasiun_name != stasiun_name:
                changes.append(f"📝 Nama stasiun: '{station.stasiun_name}' → '{stasiun_name}'")
                station.stasiun_name = stasiun_name
            
            if station.kota != kota:
                changes.append(f"📍 Kota: '{station.kota}' → '{kota}'")
                station.kota = kota
            
            # 2. Ambil data dari form
            lawan_ids = request.form.getlist('lawan_id[]')
            stasiun_lawan_list = request.form.getlist('stasiun_lawan[]')
            freq_tx_list = request.form.getlist('freq_tx[]')
            freq_rx_list = request.form.getlist('freq_rx[]')
            deleted_lawan_ids = request.form.getlist('deleted_lawan_ids[]')
            
            # Validasi minimal 1 stasiun lawan
            valid_lawan = [n for n in stasiun_lawan_list if n.strip()]
            if not valid_lawan:
                flash('❌ Minimal harus ada 1 stasiun lawan!', 'error')
                return redirect(url_for('admin_operator_edit_stasiun', station_id=station_id))
            
            # 3. Hapus lawan yang ditandai untuk dihapus
            deleted_count = 0
            if deleted_lawan_ids:
                for lawan_id in deleted_lawan_ids:
                    if lawan_id:
                        lawan = StasiunLawan.query.get(int(lawan_id))
                        if lawan and lawan.stasiun_id == station_id:
                            
                            # Putuskan relasi upload, jangan hapus file
                            uploads = UploadGambar.query.filter_by(stasiun_lawan_id=lawan.id).all()
                            for upload in uploads:
                                upload.stasiun_lawan_id = None
                            
                            # Hapus status updates
                            StatusUpdate.query.filter_by(stasiun_lawan_id=lawan.id).delete()
                            
                            changes.append(f"❌ HAPUS: '{lawan.nama_stasiun_lawan}' (status dihapus, gambar tetap ada)")
                            
                            db.session.delete(lawan)
                            deleted_count += 1
            
            # 4. Proses data yang ada (UPDATE) dan data baru (TAMBAH)
            updated_count = 0
            new_count = 0
            freq_tx_updated_count = 0
            freq_rx_updated_count = 0
            
            for i, nama_lawan in enumerate(stasiun_lawan_list):
                nama_lawan = nama_lawan.strip()
                if not nama_lawan:
                    continue
                
                freq_tx = freq_tx_list[i].strip() if i < len(freq_tx_list) else ''
                freq_rx = freq_rx_list[i].strip() if i < len(freq_rx_list) else ''
                
                # Cek apakah ini data existing (punya ID) atau baru
                if i < len(lawan_ids) and lawan_ids[i]:
                    # UPDATE DATA EXISTING
                    lawan_id = int(lawan_ids[i])
                    lawan = StasiunLawan.query.get(lawan_id)
                    
                    if lawan and lawan.stasiun_id == station_id:
                        # Cek perubahan nama
                        if lawan.nama_stasiun_lawan != nama_lawan:
                            changes.append(f"✏️ Ubah nama: '{lawan.nama_stasiun_lawan}' → '{nama_lawan}'")
                            lawan.nama_stasiun_lawan = nama_lawan
                            updated_count += 1
                        
                        # Cek perubahan frekuensi TX
                        old_freq_tx = lawan.freq_tx or ''
                        new_freq_tx = freq_tx if freq_tx else ''
                        if old_freq_tx != new_freq_tx:
                            if old_freq_tx and new_freq_tx:
                                changes.append(f"📡 Ubah Freq Tx '{lawan.nama_stasiun_lawan}': '{old_freq_tx}' → '{new_freq_tx}'")
                            elif new_freq_tx and not old_freq_tx:
                                changes.append(f"📡 Tambah Freq Tx '{lawan.nama_stasiun_lawan}': '{new_freq_tx}'")
                            elif old_freq_tx and not new_freq_tx:
                                changes.append(f"📡 Hapus Freq Tx '{lawan.nama_stasiun_lawan}' (sebelumnya '{old_freq_tx}')")
                            
                            lawan.freq_tx = new_freq_tx if new_freq_tx else None
                            freq_tx_updated_count += 1
                        
                        # Cek perubahan frekuensi RX
                        old_freq_rx = lawan.freq_rx or ''
                        new_freq_rx = freq_rx if freq_rx else ''
                        if old_freq_rx != new_freq_rx:
                            if old_freq_rx and new_freq_rx:
                                changes.append(f"📡 Ubah Freq Rx '{lawan.nama_stasiun_lawan}': '{old_freq_rx}' → '{new_freq_rx}'")
                            elif new_freq_rx and not old_freq_rx:
                                changes.append(f"📡 Tambah Freq Rx '{lawan.nama_stasiun_lawan}': '{new_freq_rx}'")
                            elif old_freq_rx and not new_freq_rx:
                                changes.append(f"📡 Hapus Freq Rx '{lawan.nama_stasiun_lawan}' (sebelumnya '{old_freq_rx}')")
                            
                            lawan.freq_rx = new_freq_rx if new_freq_rx else None
                            freq_rx_updated_count += 1
                else:
                    # DATA BARU
                    last_urutan = db.session.query(db.func.max(StasiunLawan.urutan))\
                        .filter_by(stasiun_id=station_id).scalar() or 0
                    
                    new_lawan = StasiunLawan(
                        stasiun_id=station_id,
                        nama_stasiun_lawan=nama_lawan,
                        freq_tx=freq_tx if freq_tx else None,
                        freq_rx=freq_rx if freq_rx else None,
                        group_id=None,
                        urutan=last_urutan + 1
                    )
                    db.session.add(new_lawan)
                    new_count += 1
                    freq_info = []
                    if freq_tx:
                        freq_info.append(f"Tx:{freq_tx}")
                    if freq_rx:
                        freq_info.append(f"Rx:{freq_rx}")
                    freq_str = f" ({', '.join(freq_info)})" if freq_info else ""
                    changes.append(f"✅ TAMBAH: '{nama_lawan}'{freq_str}")
            
            # 5. Commit semua perubahan
            db.session.commit()
            
            # 6. Tampilkan ringkasan perubahan
            if changes:
                flash('📋 RINCIAN PERUBAHAN:', 'info')
                for change in changes:
                    flash(change, 'debug')
            
            # Ringkasan statistik
            summary = []
            if new_count > 0:
                summary.append(f"✅ {new_count} lawan baru")
            if updated_count > 0:
                summary.append(f"✏️ {updated_count} nama diubah")
            if freq_tx_updated_count > 0:
                summary.append(f"📡 {freq_tx_updated_count} Freq Tx diubah")
            if freq_rx_updated_count > 0:
                summary.append(f"📡 {freq_rx_updated_count} Freq Rx diubah")
            if deleted_count > 0:
                summary.append(f"❌ {deleted_count} lawan dihapus")
            
            if summary:
                flash(' ✓ '.join(summary), 'success')
            else:
                flash('ℹ️ Tidak ada perubahan data.', 'info')
            
            return redirect(url_for('admin_operator_detail_stasiun', station_id=station_id))
            
        except Exception as e:
            db.session.rollback()
            error_msg = str(e)
            flash(f'❌ ERROR: {error_msg}', 'error')
            print(f"ERROR in admin_operator_edit_stasiun: {error_msg}")
            import traceback
            traceback.print_exc()
    
    return render_template('admin_operator/edit_stasiun.html',
                         station=station,
                         kota_list=KOTA_LIST)

@app.route('/admin-operator/hapus-stasiun/<int:station_id>')
@login_required
@admin_operator_required
def admin_operator_hapus_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    # Validasi kepemilikan
    if station.operator != current_user.operator_type:
        flash('❌ Stasiun ini bukan milik operator Anda!', 'error')
        return redirect(url_for('admin_operator_dashboard'))
    
    try:
        print(f"\n{'='*60}")
        print(f"🗑️  MENGHAPUS STASIUN: {station.stasiun_name} (ID: {station_id})")
        print(f"   Operator: {station.operator}")
        print(f"{'='*60}")
        
        # ===== 1. AMBIL SEMUA DATA TERKAIT =====
        # Ambil semua grup
        groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
        group_ids = [g.id for g in groups]
        print(f"   📦 Grup ditemukan: {len(group_ids)}")
        
        # Ambil semua lawan
        opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).all()
        opponent_ids = [o.id for o in opponents]
        print(f"   👥 Stasiun lawan: {len(opponent_ids)}")
        
        # Ambil semua upload
        all_uploads = UploadGambar.query.filter_by(stasiun_id=station_id).all()
        print(f"   🖼️ Upload gambar: {len(all_uploads)}")
        
        # ===== 2. HAPUS STATUS UPDATES =====
        if opponent_ids:
            deleted_status = StatusUpdate.query.filter(
                StatusUpdate.stasiun_lawan_id.in_(opponent_ids)
            ).delete(synchronize_session=False)
            print(f"   ✅ Status updates dihapus: {deleted_status}")
        
        # ===== 3. HAPUS UPLOAD GAMBAR =====
        cloudinary_deleted = 0
        for upload in all_uploads:
            try:
                # Cek apakah file masih digunakan oleh stasiun lain
                other_refs = UploadGambar.query.filter(
                    UploadGambar.public_id == upload.public_id,
                    UploadGambar.stasiun_id != station_id
                ).count()
                
                if other_refs == 0:
                    # Hapus dari Cloudinary
                    success, msg = delete_from_cloudinary(upload.public_id)
                    if success:
                        cloudinary_deleted += 1
                        print(f"      ✓ Cloudinary: {upload.public_id}")
                    else:
                        print(f"      ⚠ Cloudinary: {msg}")
                
                # Hapus record dari database
                db.session.delete(upload)
                
            except Exception as e:
                print(f"      ✗ Error hapus upload {upload.id}: {e}")
                continue
        
        print(f"   ✅ Upload gambar dihapus: {len(all_uploads)} (Cloudinary: {cloudinary_deleted})")
        
        # ===== 4. HAPUS GRUP STASIUN (URUTAN PENTING!) =====
        if group_ids:
            # PERTAMA: Putuskan relasi lawan ke grup
            for group_id in group_ids:
                updated = StasiunLawan.query.filter_by(
                    stasiun_id=station_id, 
                    grup_id=group_id
                ).update({
                    'group_id': None,
                    'grup_id': None
                })
                print(f"      ⤴️ Relasi grup {group_id} diputus: {updated} lawan")
            
            # KEDUA: Update upload yang terkait grup
            for group_id in group_ids:
                UploadGambar.query.filter_by(
                    stasiun_id=station_id,
                    group_id=group_id
                ).update({'group_id': None})
            
            # KETIGA: Hapus grup
            deleted_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).delete()
            print(f"   ✅ Grup dihapus: {deleted_groups}")
        
        # ===== 5. HAPUS STASIUN LAWAN =====
        if opponent_ids:
            deleted_opponents = StasiunLawan.query.filter_by(stasiun_id=station_id).delete()
            print(f"   ✅ Stasiun lawan dihapus: {deleted_opponents}")
        
        # ===== 6. TERAKHIR, HAPUS STASIUN =====
        db.session.delete(station)
        
        # ===== 7. COMMIT SEMUA PERUBAHAN =====
        db.session.commit()
        
        print(f"\n{'='*60}")
        print(f"✅  BERHASIL MENGHAPUS STASIUN!")
        print(f"{'='*60}")
        
        flash(f'✅ Stasiun "{station.stasiun_name}" berhasil dihapus!', 'success')
        
    except Exception as e:
        db.session.rollback()
        error_msg = f"❌ Error menghapus stasiun: {str(e)}"
        print(f"\n{'='*60}")
        print(f"❌  ERROR: {error_msg}")
        print(f"{'='*60}")
        import traceback
        traceback.print_exc()
        flash(error_msg, 'error')
    
    return redirect(url_for('admin_operator_daftar_stasiun'))

@app.route('/admin-operator/upload-excel', methods=['GET', 'POST'])
@login_required
@admin_operator_required
def admin_operator_upload_excel():
    
    operator = current_user.operator_type
    
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.files:
                flash('Tidak ada file yang dipilih!', 'error')
                return redirect(url_for('admin_operator_upload_excel'))
            
            file = request.files['excel_file']
            
            if file.filename == '':
                flash('Tidak ada file yang dipilih!', 'error')
                return redirect(url_for('admin_operator_upload_excel'))
            
            # Validasi file
            is_valid, error_msg = validate_excel_file(file)
            if not is_valid:
                flash(error_msg, 'error')
                return redirect(url_for('admin_operator_upload_excel'))
            
            kota_default = request.form.get('kota_default')
            if not kota_default:
                kota_default = None
                flash('Kota default tidak dipilih. Sistem akan menggunakan: 1) Kolom KOTA/KAB di Excel, 2) Default: Samarinda', 'info')
            
            # Simpan file sementara
            temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'temp')
            os.makedirs(temp_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_filename = f"{operator}_{timestamp}_{secure_filename(file.filename)}"
            temp_path = os.path.join(temp_dir, temp_filename)
            
            file.save(temp_path)
            
            # Process file khusus untuk operator
            success_count, error_count, messages = process_excel_upload_admin_operator(
                temp_path, 
                current_user,
                kota_default
            )
            
            # Hapus file temp
            try:
                os.remove(temp_path)
            except:
                pass
            
            # Tampilkan hasil
            if success_count > 0:
                flash(f'Berhasil menambahkan {success_count} stasiun untuk operator {operator.upper()}!', 'success')
            if error_count > 0:
                flash(f'Ada {error_count} error dalam proses upload.', 'warning')
            
            for msg in messages[-5:]:
                flash(msg, 'info')
            
            return redirect(url_for('admin_operator_daftar_stasiun'))
            
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('admin_operator_upload_excel'))
    
    # GET request
    total_stations = Stasiun.query.filter_by(operator=operator).count()
    
    total_opponents = 0
    stations = Stasiun.query.filter_by(operator=operator).all()
    for station in stations:
        total_opponents += len(station.stasiun_lawan_list)
    
    kota_list_op = db.session.query(Stasiun.kota)\
        .filter_by(operator=operator)\
        .distinct()\
        .all()
    kota_list_op = [k[0] for k in kota_list_op if k[0]]
    
    for kota in KOTA_LIST:
        if kota not in kota_list_op:
            kota_list_op.append(kota)
    
    return render_template('admin_operator/upload_excel.html',
                         operator=operator,
                         operator_name=operator.upper(),
                         total_stations=total_stations,
                         total_opponents=total_opponents,
                         kota_list=kota_list_op)

@app.route('/admin-operator/download-template')
@login_required
@admin_operator_required
def admin_operator_download_template():
    """Download template Excel untuk admin operator"""
    operator = current_user.operator_type
    return generate_admin_operator_template(operator)

# ============================================================================
# REKAP BULANAN DAN TAHUNAN
# ============================================================================

@app.route('/admin-master/rekap-bulanan')
@login_required
@admin_master_required
def admin_master_rekap_bulanan():
    """Halaman rekap bulanan ringkasan"""
    try:
        tahun = request.args.get('tahun', datetime.now().year, type=int)
        bulan = request.args.get('bulan', datetime.now().month, type=int)
        operator = request.args.get('operator', 'all')
        kota = request.args.get('kota', 'all')
        
        query = Stasiun.query
        
        if operator != 'all':
            query = query.filter(Stasiun.operator == operator)
        
        if kota != 'all':
            query = query.filter(Stasiun.kota == kota)
        
        stations = query.all()
        
        operators_list = OPERATORS
        kota_list_all = KOTA_LIST
        
        rekap_data = []
        total_aktif = 0
        total_tidak_aktif = 0
        total_tidak_berizin = 0
        total_tidak_sesuai = 0
        
        for station in stations:
            opponents = StasiunLawan.query.filter_by(stasiun_id=station.id).all()
            
            station_data = {
                'stasiun': station,
                'opponents': [],
                'total_opponents': len(opponents),
                'status_summary': {
                    'aktif': 0,
                    'tidak_aktif': 0,
                    'tidak_berizin': 0,
                    'tidak_sesuai': 0
                }
            }
            
            for opp in opponents:
                latest_status = StatusUpdate.query\
                    .filter_by(stasiun_lawan_id=opp.id)\
                    .filter(extract('year', StatusUpdate.updated_at) == tahun)\
                    .filter(extract('month', StatusUpdate.updated_at) == bulan)\
                    .order_by(StatusUpdate.updated_at.desc())\
                    .first()
                
                status = latest_status.status if latest_status else 'belum_ada'
                
                opponent_data = {
                    'id': opp.id,
                    'nama': opp.nama_stasiun_lawan,
                    'status': status,
                    'catatan': latest_status.catatan if latest_status else '',
                    'updated_at': latest_status.updated_at if latest_status else None
                }
                
                station_data['opponents'].append(opponent_data)
                
                if status in station_data['status_summary']:
                    station_data['status_summary'][status] += 1
                    
                if status == 'aktif':
                    total_aktif += 1
                elif status == 'tidak_aktif':
                    total_tidak_aktif += 1
                elif status == 'tidak_berizin':
                    total_tidak_berizin += 1
                elif status == 'tidak_sesuai':
                    total_tidak_sesuai += 1
            
            rekap_data.append(station_data)
        
        bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                     'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        return render_template('admin_master/rekap_bulanan.html',
                             rekap_data=rekap_data,
                             tahun=tahun,
                             bulan=bulan,
                             bulan_nama=bulan_nama[bulan-1],
                             operator=operator,
                             kota=kota,
                             operators=operators_list,
                             kota_list=kota_list_all,
                             total_aktif=total_aktif,
                             total_tidak_aktif=total_tidak_aktif,
                             total_tidak_berizin=total_tidak_berizin,
                             total_tidak_sesuai=total_tidak_sesuai)
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin_master_dashboard'))

@app.route('/admin-master/rekap-tahunan')
@login_required
@admin_master_required
def admin_master_rekap_tahunan():
    """Halaman rekap tahunan ringkasan"""
    try:
        tahun = request.args.get('tahun', datetime.now().year, type=int)
        operator = request.args.get('operator', 'all')
        kota = request.args.get('kota', 'all')
        
        query = Stasiun.query
        
        if operator != 'all':
            query = query.filter(Stasiun.operator == operator)
        
        if kota != 'all':
            query = query.filter(Stasiun.kota == kota)
        
        stations = query.all()
        
        operators_list = OPERATORS
        kota_list_all = KOTA_LIST
        
        bulan_nama = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                     'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        rekap_data = {}
        
        for bulan in range(1, 13):
            periode_key = f"{tahun}-{bulan:02d}"
            rekap_data[periode_key] = []
            
            for station in stations:
                opponents = StasiunLawan.query.filter_by(stasiun_id=station.id).all()
                
                station_data = {
                    'stasiun': station,
                    'opponents': [],
                    'status_summary': {
                        'aktif': 0,
                        'tidak_aktif': 0,
                        'tidak_berizin': 0,
                        'tidak_sesuai': 0
                    }
                }
                
                for opp in opponents:
                    latest_status = StatusUpdate.query\
                        .filter_by(stasiun_lawan_id=opp.id)\
                        .filter(extract('year', StatusUpdate.updated_at) == tahun)\
                        .filter(extract('month', StatusUpdate.updated_at) == bulan)\
                        .order_by(StatusUpdate.updated_at.desc())\
                        .first()
                    
                    status = latest_status.status if latest_status else 'belum_ada'
                    
                    opponent_data = {
                        'id': opp.id,
                        'nama': opp.nama_stasiun_lawan,
                        'status': status,
                        'catatan': latest_status.catatan if latest_status else '',
                        'updated_at': latest_status.updated_at if latest_status else None
                    }
                    
                    station_data['opponents'].append(opponent_data)
                    
                    if status in station_data['status_summary']:
                        station_data['status_summary'][status] += 1
                
                rekap_data[periode_key].append(station_data)
        
        return render_template('admin_master/rekap_tahunan.html',
                             rekap_data=rekap_data,
                             tahun=tahun,
                             operator=operator,
                             kota=kota,
                             operators=operators_list,
                             kota_list=kota_list_all,
                             bulan_nama=bulan_nama)
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('admin_master_dashboard'))

# ============================================================================
# USER OPERATOR ROUTES 
# ============================================================================

@app.route('/user-operator/dashboard')
@login_required
@user_operator_required
def user_operator_dashboard():
    operator = current_user.operator_type
    
    total_stations = Stasiun.query.filter_by(operator=operator).count()
    
    total_opponents = StasiunLawan.query\
        .join(Stasiun)\
        .filter(Stasiun.operator == operator)\
        .count()
    
    uploads_count = db.session.query(db.func.count(db.distinct(UploadGambar.public_id)))\
        .join(Stasiun)\
        .filter(Stasiun.operator == operator)\
        .scalar() or 0
    
    recent_uploads_subquery = db.session.query(
        UploadGambar.public_id,
        db.func.max(UploadGambar.uploaded_at).label('max_uploaded_at')
    ).join(Stasiun)\
     .filter(Stasiun.operator == operator)\
     .group_by(UploadGambar.public_id)\
     .order_by(db.desc('max_uploaded_at'))\
     .limit(5).subquery()
    
    recent_uploads = UploadGambar.query\
        .join(recent_uploads_subquery,
            db.and_(
                UploadGambar.public_id == recent_uploads_subquery.c.public_id,
                UploadGambar.uploaded_at == recent_uploads_subquery.c.max_uploaded_at
            )
        )\
        .order_by(UploadGambar.uploaded_at.desc())\
        .all()
    
    stations_without_upload = []
    all_stations = Stasiun.query.filter_by(operator=operator).all()
    for station in all_stations:
        has_upload = UploadGambar.query.filter_by(stasiun_id=station.id).first()
        if not has_upload:
            stations_without_upload.append(station)
        if len(stations_without_upload) >= 5:
            break
    
    kota_list = db.session.query(Stasiun.kota)\
        .filter_by(operator=operator)\
        .distinct()\
        .all()
    kota_list = [k[0] for k in kota_list if k[0]]
    
    current_date = datetime.now().strftime('%d %B %Y')
    
    return render_template('user_operator/dashboard.html',
                         operator=operator,
                         operator_name=operator.upper(),
                         current_date=current_date,
                         total_stations=total_stations,
                         total_uploads=uploads_count,
                         total_opponents=total_opponents,
                         recent_uploads=recent_uploads,
                         stations_without_upload=stations_without_upload,
                         kota_list=kota_list,
                         status_options=STATUS_OPTIONS)

@app.route('/user-operator/stasiun')
@login_required
@user_operator_required
def user_operator_stasiun():
    """Daftar stasiun dengan tampilan grup yang jelas"""
    operator = current_user.operator_type
    kota = request.args.get('kota', 'all')
    search_stasiun = request.args.get('search_stasiun', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    
    allowed_per_page = [10, 25, 50, 100]
    if per_page not in allowed_per_page:
        per_page = 20
    
    stations, pagination = get_stations_with_detailed_groups_paginated(
        operator, kota, search_stasiun, page, per_page
    )
    
    kota_list = db.session.query(Stasiun.kota)\
        .filter_by(operator=operator)\
        .distinct()\
        .order_by(Stasiun.kota)\
        .all()
    kota_list = [k[0] for k in kota_list if k[0]]
    
    search_params = {
        'kota': kota,
        'search_stasiun': search_stasiun,
        'per_page': per_page
    }
    
    return render_template('user_operator/stasiun.html',
                         stations=stations,
                         pagination=pagination,
                         operator=operator,
                         operator_name=operator.upper(),
                         kota_list=kota_list,
                         search_params=search_params,
                         allowed_per_page=allowed_per_page,
                         STATUS_OPTIONS=STATUS_OPTIONS)

@app.route('/user-operator/stasiun/<int:station_id>')
@login_required
@user_operator_required
def user_operator_detail_stasiun(station_id):
    """Detail stasiun dengan tampilan per grup dan status per lawan"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Stasiun ini bukan milik operator Anda!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    all_lawans = StasiunLawan.query\
        .filter_by(stasiun_id=station_id)\
        .order_by(StasiunLawan.urutan)\
        .all()
    
    # AMBIL SEMUA UPLOAD UNIK (group by public_id)
    uploads_query = db.session.query(
        UploadGambar.group_id,
        UploadGambar.public_id,
        db.func.max(UploadGambar.id).label('latest_id')
    ).filter_by(stasiun_id=station_id)\
     .group_by(UploadGambar.group_id, UploadGambar.public_id)\
     .subquery()
    
    all_uploads = UploadGambar.query\
        .join(uploads_query, UploadGambar.id == uploads_query.c.latest_id)\
        .order_by(UploadGambar.uploaded_at.desc())\
        .all()
    
    all_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
    groups_dict = {g.id: g.nama_grup for g in all_groups}
    
    groups_data = {}
    
    # === PROSES LAWAN PER GRUP ===
    for lawan in all_lawans:
        group_key = lawan.group_id if lawan.group_id is not None else 'ungrouped'
        
        if group_key not in groups_data:
            if group_key == 'ungrouped':
                label = 'Tanpa Grup'
            else:
                label = groups_dict.get(int(group_key), f'Grup {group_key}')
                
            groups_data[group_key] = {
                'group_id': lawan.group_id,
                'label': label,
                'lawans': [],
                'uploads': [],  # PASTIKAN FIELD INI ADA
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=lawan.id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        lawan_data = {
            'id': lawan.id,
            'nama': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'latest_status': latest_status.status if latest_status else None,
            'status_display': STATUS_OPTIONS.get(latest_status.status, 'Belum Ada') if latest_status else 'Belum Ada',
            'catatan': latest_status.catatan if latest_status else '',
            'updated_at': latest_status.updated_at if latest_status else None
        }
        
        groups_data[group_key]['lawans'].append(lawan_data)
        
        if latest_status and latest_status.status in groups_data[group_key]['status_summary']:
            groups_data[group_key]['status_summary'][latest_status.status] += 1
    
    # === PROSES UPLOAD PER GRUP - INI YANG DITAMBAHKAN! ===
    for upload in all_uploads:
        group_key = upload.group_id if upload.group_id is not None else 'ungrouped'
        
        if group_key not in groups_data:
            if group_key == 'ungrouped':
                label = 'Tanpa Grup'
            else:
                label = groups_dict.get(int(group_key), f'Grup {group_key}')
                
            groups_data[group_key] = {
                'group_id': upload.group_id,
                'label': label,
                'lawans': [],
                'uploads': [],
                'status_summary': {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
            }
        
        # Cek duplikasi public_id dalam grup yang sama
        existing_public_ids = [u['public_id'] for u in groups_data[group_key]['uploads']]
        if upload.public_id not in existing_public_ids:
            upload_data = {
                'id': upload.id,
                'public_id': upload.public_id,
                'cloudinary_url': upload.cloudinary_url,
                'original_filename': upload.original_filename,
                'uploaded_at': upload.uploaded_at,
                'status': upload.status,
                'format': upload.format,
                'width': upload.width,
                'height': upload.height,
                'bytes_size': upload.bytes_size
            }
            groups_data[group_key]['uploads'].append(upload_data)
    
    # Urutkan grup: ungrouped dulu, lalu grup lainnya
    sorted_groups = {}
    if 'ungrouped' in groups_data:
        sorted_groups['ungrouped'] = groups_data.pop('ungrouped')
    
    numeric_groups = {k: v for k, v in groups_data.items() if k != 'ungrouped'}
    for group_key in sorted(numeric_groups.keys(), key=lambda x: int(x) if str(x).isdigit() else x):
        sorted_groups[group_key] = numeric_groups[group_key]
    
    # Hitung status counts total
    status_counts = {'aktif': 0, 'tidak_aktif': 0, 'tidak_berizin': 0, 'tidak_sesuai': 0}
    for lawan in all_lawans:
        latest_status = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=lawan.id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        
        if latest_status and latest_status.status in status_counts:
            status_counts[latest_status.status] += 1
    
    return render_template('user_operator/detail_stasiun.html',
                         station=station,
                         groups=sorted_groups,
                         total_opponents=len(all_lawans),
                         status_counts=status_counts,
                         all_uploads=all_uploads,
                         STATUS_OPTIONS=STATUS_OPTIONS)

@app.route('/user-operator/edit-group/<int:station_id>', methods=['POST'])
@login_required
@user_operator_required
def user_operator_edit_group(station_id):
    """Edit grup"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
    
    try:
        group_id = request.form.get('group_id')
        new_group_name = request.form.get('new_group_name')
        add_opponents = request.form.getlist('add_opponents[]')
        move_opponents = request.form.getlist('move_opponents[]')
        
        for opponent_id in add_opponents:
            lawan = StasiunLawan.query.get(opponent_id)
            if lawan and lawan.stasiun_id == station_id:
                lawan.group_id = int(group_id)
        
        for opponent_id in move_opponents:
            lawan = StasiunLawan.query.get(opponent_id)
            if lawan and lawan.stasiun_id == station_id:
                new_group_key = request.form.get(f'new_group_{opponent_id}')
                if new_group_key == 'ungrouped':
                    lawan.group_id = None
                elif new_group_key:
                    lawan.group_id = int(new_group_key)
        
        db.session.commit()
        flash('Perubahan grup berhasil disimpan!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/edit-lawan-single/<int:station_id>', methods=['POST'])
@login_required
@user_operator_required
def user_operator_edit_lawan_single(station_id):
    """Edit satu stasiun lawan"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
    
    try:
        lawan_id = request.form.get('lawan_id')
        nama_lawan = request.form.get('nama_lawan')
        new_group_id = request.form.get('new_group_id')
        
        lawan = StasiunLawan.query.get(lawan_id)
        if lawan and lawan.stasiun_id == station_id:
            lawan.nama_stasiun_lawan = nama_lawan.strip()
            if new_group_id:
                lawan.group_id = int(new_group_id) if new_group_id != '' else None
        
        db.session.commit()
        flash('Data lawan berhasil diperbarui!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/delete-lawan/<int:station_id>', methods=['POST'])
@login_required
@user_operator_required
def user_operator_delete_lawan(station_id):
    """Hapus stasiun lawan (untuk yang tanpa grup)"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
    
    try:
        lawan_id = request.form.get('lawan_id')
        lawan = StasiunLawan.query.get(lawan_id)
        
        if lawan and lawan.stasiun_id == station_id:
            StatusUpdate.query.filter_by(stasiun_lawan_id=lawan.id).delete()
            
            uploads = UploadGambar.query.filter_by(stasiun_lawan_id=lawan.id).all()
            for upload in uploads:
                delete_from_cloudinary(upload.public_id)
                db.session.delete(upload)
            
            db.session.delete(lawan)
            db.session.commit()
            
            flash(f'Stasiun lawan "{lawan.nama_stasiun_lawan}" berhasil dihapus!', 'success')
        else:
            flash('Stasiun lawan tidak ditemukan', 'error')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/delete-lawan-from-group/<int:station_id>', methods=['POST'])
@login_required
@user_operator_required
def user_operator_delete_lawan_from_group(station_id):
    """
    KELUARKAN stasiun lawan dari grup - PINDAHKAN KE TANPA GRUP
    BUKAN DIHAPUS PERMANEN!
    """
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
    
    try:
        lawan_id = request.form.get('lawan_id')
        lawan = StasiunLawan.query.get(lawan_id)
        
        if lawan and lawan.stasiun_id == station_id:
            nama_lawan = lawan.nama_stasiun_lawan
            lawan.group_id = None
            
            UploadGambar.query.filter_by(
                stasiun_lawan_id=lawan.id,
                group_id=lawan.group_id
            ).update({'group_id': None})
            
            db.session.commit()
            
            flash(f'Stasiun lawan "{nama_lawan}" berhasil dikeluarkan dari grup dan dipindahkan ke "Tanpa Grup"!', 'success')
        else:
            flash('Stasiun lawan tidak ditemukan', 'error')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/grup-stasiun/<int:station_id>', methods=['GET', 'POST'])
@login_required
@user_operator_required
def user_operator_grup_stasiun(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Stasiun ini bukan milik operator Anda!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            
            # ===== 1. CREATE GROUP =====
            if action == 'create_group':
                group_name = request.form.get('group_name')
                selected_opponents = request.form.getlist('selected_opponents[]')
                
                if not group_name or not selected_opponents:
                    flash('Nama grup dan stasiun lawan harus dipilih!', 'error')
                    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
                
                new_grup = GrupStasiun(
                    stasiun_id=station_id,
                    nama_grup=group_name
                )
                db.session.add(new_grup)
                db.session.flush()
                
                for opponent_id in selected_opponents:
                    lawan = StasiunLawan.query.get(opponent_id)
                    if lawan and lawan.stasiun_id == station_id:
                        lawan.group_id = new_grup.id
                        lawan.grup_id = new_grup.id
                
                db.session.commit()
                flash(f'Grup "{group_name}" berhasil dibuat dengan {len(selected_opponents)} lawan!', 'success')
                
            # ===== 2. DELETE GROUP =====
            elif action == 'delete_group':
                group_id = request.form.get('group_id')
                if group_id:
                    grup = GrupStasiun.query.get(group_id)
                    if grup and grup.stasiun_id == station_id:
                        # Pindahkan semua lawan ke ungrouped
                        StasiunLawan.query\
                            .filter_by(stasiun_id=station_id, grup_id=group_id)\
                            .update({
                                'group_id': None,
                                'grup_id': None
                            })
                        
                        # Update upload yang terkait
                        UploadGambar.query\
                            .filter_by(stasiun_id=station_id, group_id=int(group_id))\
                            .update({'group_id': None})
                        
                        db.session.delete(grup)
                        db.session.commit()
                        flash('Grup berhasil dihapus! Semua anggota dipindahkan ke "Tanpa Grup".', 'success')
            
            # ===== 3. REMOVE FROM GROUP (KELUARKAN DARI GRUP) =====
            elif action == 'remove_from_group':
                opponent_id = request.form.get('opponent_id')
                if opponent_id:
                    lawan = StasiunLawan.query.get(opponent_id)
                    if lawan and lawan.stasiun_id == station_id:
                        nama_lawan = lawan.nama_stasiun_lawan
                        
                        # Keluarkan dari grup
                        lawan.group_id = None
                        lawan.grup_id = None
                        
                        # Update upload yang terkait
                        UploadGambar.query.filter_by(
                            stasiun_lawan_id=lawan.id
                        ).update({'group_id': None})
                        
                        db.session.commit()
                        flash(f'Stasiun lawan "{nama_lawan}" berhasil dikeluarkan dari grup!', 'success')
                    else:
                        flash('Stasiun lawan tidak ditemukan!', 'error')
            
            # ===== 4. EDIT SINGLE OPPONENT (PERBAIKAN) =====
            elif action == 'edit_lawan':
                lawan_id = request.form.get('lawan_id')
                nama_lawan = request.form.get('nama_lawan')
                freq_tx = request.form.get('freq_tx', '').strip()
                freq_rx = request.form.get('freq_rx', '').strip()
                new_group_id = request.form.get('new_group_id')
                
                # Validasi input
                if not lawan_id or not nama_lawan:
                    flash('Data tidak lengkap!', 'error')
                    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
                
                # Cari lawan
                lawan = StasiunLawan.query.get(lawan_id)
                if not lawan or lawan.stasiun_id != station_id:
                    flash('Stasiun lawan tidak ditemukan!', 'error')
                    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
                
                # Catat perubahan untuk pesan
                changes = []
                
                # Update nama
                if lawan.nama_stasiun_lawan != nama_lawan:
                    changes.append(f"Nama: '{lawan.nama_stasiun_lawan}' → '{nama_lawan}'")
                    lawan.nama_stasiun_lawan = nama_lawan
                
                # Update freq_tx (boleh kosong)
                old_freq_tx = lawan.freq_tx or ''
                new_freq_tx = freq_tx if freq_tx else None
                if old_freq_tx != (new_freq_tx or ''):
                    if old_freq_tx and new_freq_tx:
                        changes.append(f"Freq Tx: '{old_freq_tx}' → '{new_freq_tx}'")
                    elif new_freq_tx and not old_freq_tx:
                        changes.append(f"Freq Tx ditambahkan: '{new_freq_tx}'")
                    elif old_freq_tx and not new_freq_tx:
                        changes.append(f"Freq Tx dihapus (sebelumnya '{old_freq_tx}')")
                    lawan.freq_tx = new_freq_tx
                
                # Update freq_rx (boleh kosong)
                old_freq_rx = lawan.freq_rx or ''
                new_freq_rx = freq_rx if freq_rx else None
                if old_freq_rx != (new_freq_rx or ''):
                    if old_freq_rx and new_freq_rx:
                        changes.append(f"Freq Rx: '{old_freq_rx}' → '{new_freq_rx}'")
                    elif new_freq_rx and not old_freq_rx:
                        changes.append(f"Freq Rx ditambahkan: '{new_freq_rx}'")
                    elif old_freq_rx and not new_freq_rx:
                        changes.append(f"Freq Rx dihapus (sebelumnya '{old_freq_rx}')")
                    lawan.freq_rx = new_freq_rx
                
                # Update grup
                old_group = lawan.group_id
                if new_group_id and new_group_id.strip():
                    new_group_int = int(new_group_id)
                    if old_group != new_group_int:
                        # Cari nama grup baru
                        grup_baru = GrupStasiun.query.get(new_group_int)
                        grup_baru_nama = grup_baru.nama_grup if grup_baru else f'Grup {new_group_int}'
                        
                        if old_group:
                            grup_lama = GrupStasiun.query.get(old_group)
                            grup_lama_nama = grup_lama.nama_grup if grup_lama else f'Grup {old_group}'
                            changes.append(f"Grup: '{grup_lama_nama}' → '{grup_baru_nama}'")
                        else:
                            changes.append(f"Grup: 'Tanpa Grup' → '{grup_baru_nama}'")
                        
                        lawan.group_id = new_group_int
                        lawan.grup_id = new_group_int
                else:
                    if old_group is not None:
                        grup_lama = GrupStasiun.query.get(old_group)
                        grup_lama_nama = grup_lama.nama_grup if grup_lama else f'Grup {old_group}'
                        changes.append(f"Grup: '{grup_lama_nama}' → 'Tanpa Grup'")
                        lawan.group_id = None
                        lawan.grup_id = None
                
                # Update upload yang terkait dengan group_id baru
                if 'group_id' in changes or 'grup' in str(changes):
                    UploadGambar.query.filter_by(
                        stasiun_lawan_id=lawan.id
                    ).update({'group_id': lawan.group_id})
                
                # Commit perubahan
                db.session.commit()
                
                # Tampilkan pesan
                if changes:
                    flash(f'✅ Stasiun lawan "{nama_lawan}" berhasil diperbarui!', 'success')
                    for change in changes:
                        flash(f'  • {change}', 'info')
                else:
                    flash(f'ℹ️ Tidak ada perubahan pada "{nama_lawan}".', 'info')
            
            return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'❌ Error: {str(e)}', 'error')
            print(f"ERROR in user_operator_grup_stasiun: {str(e)}")
            import traceback
            traceback.print_exc()
            return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))
    
    # ===== GET REQUEST - TAMPILKAN HALAMAN =====
    stasiun_lawan = StasiunLawan.query\
        .filter_by(stasiun_id=station_id)\
        .order_by(StasiunLawan.urutan)\
        .all()
    
    all_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
    groups_dict = {g.id: g for g in all_groups}
    
    groups_with_names = {}
    
    for lawan in stasiun_lawan:
        if lawan.grup_id:
            group_key = lawan.grup_id
        else:
            group_key = 'ungrouped'
        
        if group_key not in groups_with_names:
            if group_key == 'ungrouped':
                groups_with_names[group_key] = {
                    'nama': 'Tanpa Grup',
                    'lawans': []
                }
            else:
                grup_obj = groups_dict.get(group_key)
                groups_with_names[group_key] = {
                    'nama': grup_obj.nama_grup if grup_obj else f'Grup {group_key}',
                    'lawans': []
                }
        
        lawan_data = {
            'id': lawan.id,
            'nama_stasiun_lawan': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'group_id': lawan.group_id
        }
        groups_with_names[group_key]['lawans'].append(lawan_data)
    
    ungrouped_lawans = StasiunLawan.query\
        .filter_by(stasiun_id=station_id, grup_id=None)\
        .order_by(StasiunLawan.urutan)\
        .all()
    
    return render_template('user_operator/grup_stasiun.html',
                         station=station,
                         stasiun_lawan=stasiun_lawan,
                         groups=groups_with_names,
                         ungrouped_lawans=ungrouped_lawans)

@app.route('/user-operator/upload/<int:station_id>', methods=['GET', 'POST'])
@login_required
@user_operator_required
def user_operator_upload(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    # GET request
    if request.method == 'GET':
        all_groups = GrupStasiun.query.filter_by(stasiun_id=station_id).all()
        station.grup_list = all_groups
        
        stasiun_lawan = StasiunLawan.query\
            .filter_by(stasiun_id=station_id)\
            .order_by(StasiunLawan.urutan)\
            .all()
        
        groups = {}
        for lawan in stasiun_lawan:
            group_key = lawan.group_id if lawan.group_id is not None else 'ungrouped'
            if group_key not in groups:
                groups[group_key] = []
            
            latest_status = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=lawan.id)\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
            
            lawan_data = {
                'id': lawan.id,
                'nama': lawan.nama_stasiun_lawan,
                'freq_tx': lawan.freq_tx,
                'freq_rx': lawan.freq_rx,
                'latest_status': latest_status,
                'catatan': latest_status.catatan if latest_status else ''
            }
            groups[group_key].append(lawan_data)
        
        return render_template('user_operator/upload.html',
                             station=station,
                             groups=groups,
                             status_options=STATUS_OPTIONS)
    
    # POST REQUEST - UPDATE KE CLOUDINARY
    if request.method == 'POST':
        try:
            print("=== UPLOAD TO CLOUDINARY START ===")
            
            # Koleksi semua file yang diupload
            saved_files = []
            
            # Proses setiap file dari form
            for key in request.files:
                files = request.files.getlist(key)
                
                for file in files:
                    if file and file.filename:
                        print(f"Processing: {file.filename}")
                        
                        # Tentukan group_id dari key form
                        if key.startswith('gambar_'):
                            group_key = key.replace('gambar_', '')
                        else:
                            group_key = 'ungrouped'
                        
                        group_id = None
                        if group_key != 'ungrouped':
                            try:
                                group_id = int(group_key)
                            except:
                                group_id = None
                        
                        # Upload ke Cloudinary
                        result, error = upload_to_cloudinary(
                            file, 
                            station_id, 
                            None, 
                            group_id
                        )
                        
                        if result:
                            saved_files.append({
                                'public_id': result['public_id'],
                                'url': result['url'],
                                'format': result['format'],
                                'width': result['width'],
                                'height': result['height'],
                                'bytes': result['bytes'],
                                'original_filename': result['original_filename'],
                                'group_id': group_id,
                                'group_key': group_key
                            })
                            print(f"✅ Uploaded: {result['public_id']}")
                        else:
                            flash(f"Gagal upload {file.filename}: {error}", 'error')
            
            # Proses status updates dari form
            status_updates = []
            for key, value in request.form.items():
                if key.startswith('status_'):
                    parts = key.split('_')
                    if len(parts) >= 3:
                        group_key = parts[1]
                        lawan_id = parts[-1]
                        
                        catatan_key = f"catatan_{group_key}_{lawan_id}"
                        catatan = request.form.get(catatan_key, '')
                        
                        if value:
                            status_updates.append({
                                'lawan_id': int(lawan_id),
                                'status': value,
                                'catatan': catatan,
                                'group_key': group_key
                            })
            
            print(f"Status updates to process: {len(status_updates)}")
            
            upload_count = 0
            status_count = 0
            
            for status_update in status_updates:
                lawan_id = status_update['lawan_id']
                status = status_update['status']
                catatan = status_update['catatan']
                group_key = status_update['group_key']
                
                try:
                    # 1. Simpan status update
                    new_status = StatusUpdate(
                        stasiun_lawan_id=lawan_id,
                        status=status,
                        updated_by=current_user.id,
                        catatan=catatan
                    )
                    db.session.add(new_status)
                    status_count += 1
                    
                    # 2. Simpan upload records (untuk setiap file yang sesuai grup)
                    for saved_file in saved_files:
                        if saved_file['group_key'] == group_key:
                            upload = UploadGambar(
                                stasiun_id=station_id,
                                stasiun_lawan_id=lawan_id,
                                group_id=saved_file['group_id'],
                                # Data Cloudinary
                                public_id=saved_file['public_id'],
                                cloudinary_url=saved_file['url'],
                                original_filename=saved_file['original_filename'],
                                width=saved_file['width'],
                                height=saved_file['height'],
                                format=saved_file['format'],
                                bytes_size=saved_file['bytes'],
                                # Data lain
                                status=status,
                                uploaded_by=current_user.id,
                                is_checked=True
                            )
                            db.session.add(upload)
                            upload_count += 1
                            print(f"✓ Created upload for lawan {lawan_id}: {saved_file['public_id']}")
                            
                except Exception as e:
                    print(f"✗ Error saving for lawan {lawan_id}: {e}")
                    continue
            
            db.session.commit()
            
            if upload_count > 0 or status_count > 0:
                message = f"✅ Berhasil! "
                if upload_count > 0:
                     " Tambah gambar, "
                if status_count > 0:
                    message += f"{status_count} status diperbarui."
                flash(message, 'success')
            else:
                flash('Tidak ada data yang disimpan.', 'info')
            
            print(f"=== UPLOAD COMPLETE ===")
            print(f"Upload records: {upload_count}")
            print(f"Status updates: {status_count}")
            
            return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))
            
        except Exception as e:
            db.session.rollback()
            print(f"=== UPLOAD ERROR ===")
            print(f"Error: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Error: {str(e)}', 'error')
            return redirect(url_for('user_operator_upload', station_id=station_id))

@app.route('/user-operator/delete-opponent/<int:station_id>/<int:opponent_id>', methods=['POST'])
@login_required
@user_operator_required
def user_operator_delete_opponent(station_id, opponent_id):
    """Hapus stasiun lawan dari database"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    try:
        lawan = StasiunLawan.query.get_or_404(opponent_id)
        
        StatusUpdate.query.filter_by(stasiun_lawan_id=opponent_id).delete()
        
        uploads = UploadGambar.query.filter_by(stasiun_lawan_id=opponent_id).all()
        for upload in uploads:
            delete_from_cloudinary(upload.public_id)
            db.session.delete(upload)
        
        db.session.delete(lawan)
        db.session.commit()
        
        flash('Stasiun lawan berhasil dihapus!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/edit-grup/<int:station_id>/<group_key>', methods=['GET', 'POST'])
@login_required
@user_operator_required
def user_operator_edit_grup(station_id, group_key):
    """Edit grup - UPDATE UNTUK CLOUDINARY"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    if group_key == 'ungrouped':
        group_id = None
        group_label = 'Tanpa Grup'
    else:
        try:
            group_id = int(group_key)
            grup = GrupStasiun.query.filter_by(id=group_id, stasiun_id=station_id).first()
            if grup:
                group_label = grup.nama_grup
            else:
                group_label = f'Grup {group_id}'
        except:
            flash('Group key tidak valid!', 'error')
            return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))
    
    if group_id is None:
        lawans_in_group = StasiunLawan.query.filter_by(stasiun_id=station_id, group_id=None).all()
    else:
        lawans_in_group = StasiunLawan.query.filter_by(stasiun_id=station_id, group_id=group_id).all()
    
    if group_id is None:
        uploads_sub = db.session.query(
            UploadGambar.public_id, 
            db.func.max(UploadGambar.id).label('id')
        ).filter(
            UploadGambar.stasiun_id == station_id, 
            UploadGambar.group_id.is_(None)
        ).group_by(UploadGambar.public_id).subquery()
    else:
        uploads_sub = db.session.query(
            UploadGambar.public_id, 
            db.func.max(UploadGambar.id).label('id')
        ).filter_by(
            stasiun_id=station_id, 
            group_id=group_id
        ).group_by(UploadGambar.public_id).subquery()
    
    current_uploads = UploadGambar.query.join(
        uploads_sub, UploadGambar.id == uploads_sub.c.id
    ).order_by(UploadGambar.uploaded_at.desc()).all()
    
    if request.method == 'POST':
        try:
            print("\n" + "="*80)
            print("🚀 EDIT GRUP - CLOUDINARY")
            print("="*80)
            
            # ===== DEBUG: CEK SEMUA FILE DI REQUEST =====
            print(f"\n📋 REQUEST FILES:")
            print(f"  request.files keys: {list(request.files.keys())}")
            
            files = request.files.getlist('new_files[]')
            print(f"  files from 'new_files[]': {len(files)}")
            
            for i, file in enumerate(files):
                if file and file.filename:
                    print(f"    File {i}: {file.filename} - {file.content_type} - {file.content_length} bytes")
            
            # ===== DEBUG: CEK SEMUA FORM DATA =====
            print(f"\n📋 FORM DATA:")
            for key, value in request.form.items():
                if not key.startswith('status_') and not key.startswith('catatan_'):
                    print(f"  {key}: {value}")
            
            delete_files = request.form.getlist('delete_files[]')
            print(f"\n🗑️ Files to delete: {delete_files}")
            
            # ===== 1. HAPUS FILE =====
            deleted_count = 0
            for upload_id in delete_files:
                try:
                    upload_id_int = int(upload_id)
                    upload = UploadGambar.query.get(upload_id_int)
                    
                    if upload and upload.stasiun_id == station_id:
                        public_id = upload.public_id
                        print(f"  Deleting: {public_id}")
                        
                        # Cari semua record dengan public_id yang sama dalam grup ini
                        deleted_records = UploadGambar.query.filter_by(
                            stasiun_id=station_id,
                            public_id=public_id,
                            group_id=group_id
                        ).delete(synchronize_session=False)
                        
                        print(f"    - {deleted_records} records deleted")
                        
                        # Cek apakah masih ada record lain dengan public_id yang sama
                        other_refs = UploadGambar.query.filter_by(public_id=public_id).count()
                        if other_refs == 0:
                            success, msg = delete_from_cloudinary(public_id)
                            if success:
                                print(f"    - Cloudinary file deleted")
                        
                        deleted_count += deleted_records
                except Exception as e:
                    print(f"    - Error: {e}")
            
            # ===== 2. KOLEKSI STATUS DARI FORM =====
            status_map = {}
            for lawan in lawans_in_group:
                status_value = request.form.get(f'status_{lawan.id}')
                catatan_value = request.form.get(f'catatan_{lawan.id}', '')
                if status_value:
                    status_map[lawan.id] = {
                        'status': status_value,
                        'catatan': catatan_value
                    }
                    print(f"  Status for lawan {lawan.id}: {status_value}")
            
            # ===== 3. UPLOAD FILE BARU - CARA ALTERNATIF =====
            uploaded_count = 0
            
            # Coba cari file dengan cara langsung
            for key in request.files:
                file_list = request.files.getlist(key)
                print(f"  Processing key '{key}' with {len(file_list)} files")
                
                for file in file_list:
                    if file and file.filename:
                        print(f"\n  📤 Processing: {file.filename}")
                        
                        # Validasi file
                        if not allowed_file(file.filename):
                            print(f"    - ❌ Format tidak didukung: {file.filename}")
                            flash(f'File {file.filename} format tidak didukung!', 'warning')
                            continue
                        
                        # Baca file untuk validasi ukuran
                        file_content = file.read()
                        file_size = len(file_content)
                        file.seek(0)  # Reset pointer
                        
                        if file_size > 5 * 1024 * 1024:
                            print(f"    - ❌ File terlalu besar: {file_size} bytes")
                            flash(f'File {file.filename} terlalu besar (maks 5MB)!', 'warning')
                            continue
                        
                        print(f"    - ✅ File valid: {file.filename}, size: {file_size} bytes")
                        
                        # Upload ke Cloudinary
                        result, error = upload_to_cloudinary(file, station_id, None, group_id)
                        
                        if result:
                            print(f"    - ✅ File uploaded ke Cloudinary: {result['public_id']}")
                            
                            # Untuk setiap lawan dalam grup, buat record upload
                            for lawan in lawans_in_group:
                                # Ambil status lawan ini dari form
                                status_info = status_map.get(lawan.id, {})
                                current_status = status_info.get('status')
                                current_catatan = status_info.get('catatan', '')
                                
                                upload = UploadGambar(
                                    stasiun_id=station_id,
                                    stasiun_lawan_id=lawan.id,
                                    group_id=group_id,
                                    # Data Cloudinary
                                    public_id=result['public_id'],
                                    cloudinary_url=result['url'],
                                    original_filename=result['original_filename'],
                                    width=result['width'],
                                    height=result['height'],
                                    format=result['format'],
                                    bytes_size=result['bytes'],
                                    # Data lain
                                    status=current_status,
                                    uploaded_by=current_user.id,
                                    is_checked=True
                                )
                                db.session.add(upload)
                                uploaded_count += 1
                                print(f"    - Created upload for lawan {lawan.id} with status '{current_status}'")
                        else:
                            print(f"    - ❌ Failed to upload: {error}")
                            flash(f'Gagal upload {file.filename}: {error}', 'error')
            
            # ===== 4. UPDATE STATUS =====
            status_count = 0
            for lawan in lawans_in_group:
                status_info = status_map.get(lawan.id)
                
                if status_info:
                    print(f"\n  Updating status for lawan {lawan.id}: {status_info['status']}")
                    
                    # Buat status update baru
                    status_update = StatusUpdate(
                        stasiun_lawan_id=lawan.id,
                        status=status_info['status'],
                        updated_by=current_user.id,
                        catatan=status_info['catatan']
                    )
                    db.session.add(status_update)
                    status_count += 1
                    
                    # Update status di SEMUA upload yang terkait dengan lawan ini
                    updated = UploadGambar.query.filter_by(
                        stasiun_lawan_id=lawan.id
                    ).update({'status': status_info['status']})
                    print(f"    - Updated {updated} existing uploads")
            
            # ===== 5. COMMIT SEMUA PERUBAHAN =====
            db.session.commit()
            
            print("\n" + "="*80)
            print(f"✅ SUMMARY:")
            print(f"  - Deleted: {deleted_count} records")
            print(f"  - Uploaded: {uploaded_count} files")
            print(f"  - Status updated: {status_count}")
            print("="*80)
            
            if status_count > 0:
                flash(f'✅ {status_count} status diperbarui', 'success')
            
        except Exception as e:
            db.session.rollback()
            print(f"\n❌ ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Error: {str(e)}', 'error')
        
        return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))
    
    # GET REQUEST - TAMPILKAN FORM
    opponents_data = []
    for lawan in lawans_in_group:
        latest = StatusUpdate.query.filter_by(
            stasiun_lawan_id=lawan.id
        ).order_by(StatusUpdate.updated_at.desc()).first()
        
        opponents_data.append({
            'id': lawan.id,
            'nama': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'current_status': latest.status if latest else None,
            'status_display': STATUS_OPTIONS.get(latest.status, 'Belum Ada') if latest else 'Belum Ada',
            'catatan': latest.catatan if latest else ''
        })
    
    return render_template('user_operator/edit_grup.html',
                         station=station,
                         group_key=group_key,
                         group_label=group_label,
                         group_id=group_id,
                         current_uploads=current_uploads,
                         opponents=opponents_data,
                         status_options=STATUS_OPTIONS)

@app.route('/user-operator/hapus-upload/<int:upload_id>')
@login_required
@user_operator_required
def user_operator_hapus_upload(upload_id):
    """Hapus upload dan semua duplikatnya dari Cloudinary"""
    upload = UploadGambar.query.get_or_404(upload_id)
    station = Stasiun.query.get(upload.stasiun_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    try:
        # Cari semua record dengan public_id yang sama
        same_file_uploads = UploadGambar.query.filter_by(
            stasiun_id=upload.stasiun_id,
            public_id=upload.public_id,
            group_id=upload.group_id
        ).all()
        
        # Cek apakah ada record lain dengan public_id yang sama (untuk stasiun lain)
        other_uploads = UploadGambar.query\
            .filter(UploadGambar.public_id == upload.public_id,
                   UploadGambar.id != upload.id).count()
        
        # Hapus dari Cloudinary (hanya jika ini adalah satu-satunya record)
        if other_uploads == 0:
            success, message = delete_from_cloudinary(upload.public_id)
            if success:
                print(f"✅ Deleted from Cloudinary: {upload.public_id}")
            else:
                print(f"⚠️ Cloudinary delete issue: {message}")
        
        # Hapus record dari database
        for u in same_file_uploads:
            db.session.delete(u)
        
        db.session.commit()
        
        flash(f'✅ Berhasil menghapus {len(same_file_uploads)} gambar dari Cloudinary!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'❌ Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_detail_stasiun', station_id=upload.stasiun_id))

@app.route('/user-operator/cleanup-files/<int:station_id>')
@login_required
@user_operator_required
def user_operator_cleanup_files(station_id):
    """Bersihkan file fisik yang tidak terpakai - TIDAK PERLU UNTUK CLOUDINARY"""
    flash('Fitur cleanup tidak diperlukan karena menggunakan Cloudinary', 'info')
    return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))

@app.route('/api/delete-group-uploads/<int:station_id>', methods=['POST'])
@login_required
@user_operator_required
def api_delete_group_uploads(station_id):
    """Delete semua upload dalam satu grup dari Cloudinary"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        return jsonify({'success': False, 'message': 'Akses ditolak!'}), 403
    
    try:
        data = request.get_json()
        group_id = data.get('group_id')
        
        if not group_id:
            return jsonify({'success': False, 'message': 'Group ID harus diisi!'})
        
        # Ambil semua upload dalam grup
        if group_id == 'ungrouped':
            uploads = UploadGambar.query.filter(
                UploadGambar.stasiun_id == station_id,
                UploadGambar.group_id.is_(None)
            ).all()
        else:
            uploads = UploadGambar.query.filter_by(
                stasiun_id=station_id,
                group_id=group_id
            ).all()
        
        # Koleksi public_id unik
        unique_public_ids = set()
        deleted_count = 0
        
        # Hapus dari Cloudinary (sekali per public_id unik)
        for upload in uploads:
            if upload.public_id not in unique_public_ids:
                # Cek apakah public_id ini dipakai di tempat lain
                other_refs = UploadGambar.query\
                    .filter(UploadGambar.public_id == upload.public_id,
                           UploadGambar.id != upload.id).count()
                
                if other_refs == 0:
                    delete_from_cloudinary(upload.public_id)
                
                unique_public_ids.add(upload.public_id)
            
            db.session.delete(upload)
            deleted_count += 1
        
        # Hapus status updates untuk lawan dalam grup
        if group_id == 'ungrouped':
            for upload in uploads:
                if upload.stasiun_lawan_id:
                    StatusUpdate.query.filter_by(stasiun_lawan_id=upload.stasiun_lawan_id).delete()
        else:
            lawans_in_group = StasiunLawan.query.filter_by(
                stasiun_id=station_id,
                group_id=group_id
            ).all()
            
            for lawan in lawans_in_group:
                StatusUpdate.query.filter_by(stasiun_lawan_id=lawan.id).delete()
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'{deleted_count} gambar dalam grup berhasil dihapus dari Cloudinary!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/user-operator/fix-group-ids/<int:station_id>')
@login_required
@user_operator_required
def user_operator_fix_group_ids(station_id):
    """Route untuk memperbaiki group_id di upload_gambar"""
    station = Stasiun.query.get_or_404(station_id)
    
    if station.operator != current_user.operator_type:
        flash('Akses ditolak!', 'error')
        return redirect(url_for('user_operator_dashboard'))
    
    try:
        uploads = UploadGambar.query.filter_by(stasiun_id=station_id).all()
        
        for upload in uploads:
            if upload.stasiun_lawan_id:
                lawan = StasiunLawan.query.get(upload.stasiun_lawan_id)
                if lawan:
                    upload.group_id = lawan.group_id
        
        db.session.commit()
        flash('Group IDs berhasil diperbaiki!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))

# ============================================================================
# FITUR LAMA (UNTUK KOMPATIBILITAS)
# ============================================================================

@app.route('/user-operator/manage-groups/<int:station_id>')
@login_required
@user_operator_required
def user_operator_manage_groups(station_id):
    return redirect(url_for('user_operator_grup_stasiun', station_id=station_id))

@app.route('/user-operator/upload-gambar/<int:station_id>', methods=['GET', 'POST'])
@login_required
@user_operator_required
def user_operator_upload_gambar(station_id):
    return redirect(url_for('user_operator_upload', station_id=station_id))

@app.route('/user-operator/manage-uploads/<int:station_id>')
@login_required
@user_operator_required
def user_operator_manage_uploads(station_id):
    return redirect(url_for('user_operator_detail_stasiun', station_id=station_id))

@app.route('/user-operator/card-view')
@login_required
@user_operator_required
def user_operator_card_view():
    return redirect(url_for('user_operator_stasiun'))

# ============================================================================
# API ENDPOINTS - UPDATE UNTUK CLOUDINARY
# ============================================================================

@app.route('/api/upload/<int:upload_id>')
@login_required
def api_get_upload_detail(upload_id):
    upload = UploadGambar.query.get_or_404(upload_id)
    
    station = Stasiun.query.get(upload.stasiun_id)
    if current_user.role == 'user_operator' and station.operator != current_user.operator_type:
        return jsonify({'error': 'Akses ditolak'}), 403
    
    catatan = ''
    if upload.stasiun_lawan_id:
        status_update = StatusUpdate.query\
            .filter_by(stasiun_lawan_id=upload.stasiun_lawan_id)\
            .order_by(StatusUpdate.updated_at.desc())\
            .first()
        if status_update:
            catatan = status_update.catatan or ''
    
    # Dapatkan thumbnail URL
    thumbnail_url = None
    if upload.public_id:
        thumbnail_url, _ = cloudinary_url(
            upload.public_id, 
            width=200, 
            height=200, 
            crop="fill", 
            quality="auto"
        )
    
    return jsonify({
        'success': True,
        'upload': {
            'id': upload.id,
            'public_id': upload.public_id,
            'url': upload.cloudinary_url,
            'thumbnail_url': thumbnail_url,
            'original_filename': upload.original_filename,
            'status': upload.status,
            'width': upload.width,
            'height': upload.height,
            'format': upload.format,
            'size_mb': round(upload.bytes_size / (1024 * 1024), 2) if upload.bytes_size else None,
            'uploaded_at': upload.uploaded_at.strftime('%Y-%m-%d %H:%M:%S')
        },
        'catatan': catatan
    })

@app.route('/api/update_upload_status/<int:upload_id>', methods=['PUT'])
@login_required
def api_update_upload_status(upload_id):
    upload = UploadGambar.query.get_or_404(upload_id)
    
    station = Stasiun.query.get(upload.stasiun_id)
    if current_user.role == 'user_operator' and station.operator != current_user.operator_type:
        return jsonify({'error': 'Akses ditolak'}), 403
    
    try:
        data = request.get_json()
        status = data.get('status')
        catatan = data.get('catatan', '')
        
        if not status:
            return jsonify({'success': False, 'message': 'Status harus diisi!'})
        
        upload.status = status
        
        if upload.stasiun_lawan_id:
            status_update = StatusUpdate.query\
                .filter_by(stasiun_lawan_id=upload.stasiun_lawan_id)\
                .order_by(StatusUpdate.updated_at.desc())\
                .first()
            
            if not status_update:
                status_update = StatusUpdate(
                    stasiun_lawan_id=upload.stasiun_lawan_id,
                    status=status,
                    updated_by=current_user.id,
                    catatan=catatan
                )
                db.session.add(status_update)
            else:
                status_update.status = status
                status_update.updated_by = current_user.id
                status_update.catatan = catatan
                status_update.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Status berhasil diperbarui!'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/delete_upload/<int:upload_id>', methods=['DELETE'])
@login_required
def api_delete_upload(upload_id):
    upload = UploadGambar.query.get_or_404(upload_id)
    
    station = Stasiun.query.get(upload.stasiun_id)
    if current_user.role == 'user_operator' and station.operator != current_user.operator_type:
        return jsonify({'error': 'Akses ditolak'}), 403
    
    try:
        # Hapus dari Cloudinary
        success, message = delete_from_cloudinary(upload.public_id)
        
        # Hapus dari database
        db.session.delete(upload)
        db.session.commit()
        
        if success:
            return jsonify({
                'success': True,
                'message': 'Gambar berhasil dihapus dari Cloudinary!'
            })
        else:
            return jsonify({
                'success': True,
                'message': f'Gambar dihapus dari database, tapi mungkin masih di Cloudinary: {message}'
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/stasiun/<int:station_id>/uploads-detail')
@login_required
def api_get_uploads_detail(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if current_user.role == 'user_operator' and station.operator != current_user.operator_type:
        return jsonify({'error': 'Akses ditolak'}), 403
    
    uploads = UploadGambar.query.filter_by(stasiun_id=station_id).all()
    
    result = []
    for upload in uploads:
        opponent_name = None
        if upload.stasiun_lawan_id:
            lawan = StasiunLawan.query.get(upload.stasiun_lawan_id)
            opponent_name = lawan.nama_stasiun_lawan if lawan else None
        
        # Dapatkan thumbnail URL
        thumbnail_url = None
        if upload.public_id:
            thumbnail_url, _ = cloudinary_url(
                upload.public_id, 
                width=200, 
                height=200, 
                crop="fill", 
                quality="auto"
            )
        
        result.append({
            'id': upload.id,
            'public_id': upload.public_id,
            'url': upload.cloudinary_url,
            'thumbnail_url': thumbnail_url,
            'original_filename': upload.original_filename,
            'status': upload.status,
            'opponent': opponent_name,
            'date': upload.uploaded_at.strftime('%d/%m/%Y'),
            'format': upload.format,
            'width': upload.width,
            'height': upload.height,
            'size_mb': round(upload.bytes_size / (1024 * 1024), 2) if upload.bytes_size else None
        })
    
    return jsonify(result)

@app.route('/api/stasiun-lawan/<int:station_id>/groups')
@login_required
def api_get_stasiun_lawan_groups(station_id):
    station = Stasiun.query.get_or_404(station_id)
    
    if current_user.role == 'user_operator' and station.operator != current_user.operator_type:
        return jsonify({'error': 'Akses ditolak'}), 403
    
    stasiun_lawan = StasiunLawan.query.filter_by(stasiun_id=station_id).order_by(StasiunLawan.urutan).all()
    
    groups = {}
    for lawan in stasiun_lawan:
        group_key = lawan.group_id if lawan.group_id is not None else 'no_group'
        if group_key not in groups:
            groups[group_key] = {
                'group_id': lawan.group_id,
                'opponents': []
            }
        groups[group_key]['opponents'].append({
            'id': lawan.id,
            'nama': lawan.nama_stasiun_lawan,
            'freq_tx': lawan.freq_tx,
            'freq_rx': lawan.freq_rx,
            'group_id': lawan.group_id,
            'urutan': lawan.urutan
        })
    
    result = []
    for group_key, group_data in groups.items():
        result.append(group_data)
    
    return jsonify(result)

# ============================================================================
# ROUTE UNTUK MELIHAT GAMBAR (TIDAK DIGUNAKAN LAGI)
# ============================================================================

@app.route('/uploads/<path:filename>')
def serve_uploaded_file(filename):
    """Tidak digunakan lagi karena pakai Cloudinary"""
    return "Gambar menggunakan Cloudinary", 404

# ============================================================================
# INITIALIZATION
# ============================================================================

def init_database():
    """Inisialisasi database dengan data default"""
    with app.app_context():
        try:
            db.create_all()
            
            if User.query.count() == 0:
                users_data = [
                    {
                        'username': 'admin_master',
                        'password': 'adminmasterbalmon2026',
                        'role': 'admin_master',
                        'operator_type': None
                    },
                    {
                        'username': 'pic_telkom',
                        'password': 'pictelkombalmon2026',
                        'role': 'admin_operator',
                        'operator_type': 'telkom'
                    },
                    {
                        'username': 'pic_telkomsel',
                        'password': 'pictelkomselbalmon2026',
                        'role': 'admin_operator',
                        'operator_type': 'telkomsel'
                    },
                    {
                        'username': 'pic_indosat',
                        'password': 'picindosatbalmon2026',
                        'role': 'admin_operator',
                        'operator_type': 'indosat'
                    },
                    {
                        'username': 'pic_xl',
                        'password': 'picxlbalmon2026',
                        'role': 'admin_operator',
                        'operator_type': 'xl'
                    },
                    {
                        'username': 'telkom',
                        'password': 'operatortelkombalmon01',
                        'role': 'user_operator',
                        'operator_type': 'telkom'
                    },
                    {
                        'username': 'telkomsel',
                        'password': 'operatortelkomselbalmon02',
                        'role': 'user_operator',
                        'operator_type': 'telkomsel'
                    },
                    {
                        'username': 'indosat',
                        'password': 'operatorindosatbalmon010',
                        'role': 'user_operator',
                        'operator_type': 'indosat'
                    },
                    {
                        'username': 'xl',
                        'password': 'operatorxlbalmon20',
                        'role': 'user_operator',
                        'operator_type': 'xl'
                    }
                ]
                
                for user_data in users_data:
                    hashed_password = generate_password_hash(user_data['password'], method='pbkdf2:sha256')
                    user = User(
                        username=user_data['username'],
                        password=hashed_password,
                        role=user_data['role'],
                        operator_type=user_data['operator_type']
                    )
                    db.session.add(user)
                
                db.session.commit()
                print("=" * 60)
                print("DATABASE BERHASIL DIINISIALISASI")
                print("=" * 60)
                print("User login telah dibuat sesuai dengan data di database.")
                print("Silakan gunakan kredensial yang telah diberikan.")
                print("=" * 60)
            else:
                print("Database sudah ada, melewati inisialisasi...")
            
            create_upload_folders()
            print("Folder upload berhasil dibuat/dicek.")
            print("=" * 60)
                
        except Exception as e:
            print(f"Error initializing database: {e}")
            db.session.rollback()

# ============================================================================
# MAIN EXECUTION
# ============================================================================

if __name__ == '__main__':
    print("Starting Station Monitoring System...")
    print("Initializing database...")
    init_database()
    print("Application is running on http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)